"""宅建士の公式様式とパイプライン出力をセル単位で照合する QA ツール.

公式様式WB（人手で仕上げた正解）に対し、同一物件の構造化データを
パイプライン（build_juyojiko / build_keiyaku → wb_fill）へ通した結果が
どれだけ一致するかをセル単位で測る。座標マッピングの回帰確認に使う。

指標:
  - 一致率   : パイプラインが値を入れたセルのうち、公式様式と一致した割合
  - カバー率 : 宅建士が記入した案件固有セル（公式様式≠ブランク）のうち、
               パイプラインが充当した割合
  - gaps     : 宅建士が記入したのにパイプラインが未充当のセル（手作業が残る箇所）

PII を含む実物件WBはリポジトリに含めない。外部パスで渡す。

CLI::
    python fidelity_check.py 構造化.json 36-1 ブランク.xlsx 公式様式.xlsx --doc juyojiko

構造化.json は Juyojiko（--doc juyojiko）/ Keiyakusho（--doc keiyaku）の
model_dump 互換 JSON。
"""

from __future__ import annotations

import io
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import cellmaps
import wb_fill
from juyojiko_schema import Juyojiko
from keiyaku_schema import Keiyakusho

SHEETS = {"juyojiko": cellmaps.JUYOJIKO_SHEET, "keiyaku": cellmaps.CONTRACT_SHEET}


def _norm(v: Any) -> str:
    """比較用の正規化（前後空白・全角/半角空白を無視）。"""
    if v in (None, ""):
        return ""
    return str(v).replace(" ", "").replace("　", "").strip()


def _fill(bc: Any, variant: str, doc: str, blank_path: str) -> Any:
    """ブランク様式へパイプライン差込し、対象シートを返す（差込値 dict も）。"""
    if doc == "juyojiko":
        sv, sc = cellmaps.build_juyojiko(variant, bc)
    else:
        sv, sc = cellmaps.build_keiyaku(variant, bc)
    out, _ = wb_fill.fill_workbook(open(blank_path, "rb").read(), sv, sc)
    sheet = SHEETS[doc]
    return load_workbook(io.BytesIO(out))[sheet], sv[sheet]


def compare(bc: Any, variant: str, doc: str, blank_path: str,
            truth_path: str) -> dict[str, Any]:
    """パイプライン出力と公式様式を照合し、指標と差分・未カバーを返す。"""
    sheet = SHEETS[doc]
    _, written_all = _fill(bc, variant, doc, blank_path)
    written = {c: v for c, v in written_all.items() if v not in (None, "")}
    truth = load_workbook(truth_path, data_only=True)[sheet]
    blank = load_workbook(blank_path, data_only=True)[sheet]

    match, diffs = 0, []
    for coord, pv in written.items():
        r, c = coordinate_to_tuple(coord)
        tv = truth.cell(r, c).value
        if _norm(pv) == _norm(tv):
            match += 1
        else:
            diffs.append((coord, pv, tv))

    # 宅建士が記入した案件固有セル = 公式様式が非空かつブランクと異なる
    human: dict[str, Any] = {}
    for r in range(1, truth.max_row + 1):
        for c in range(1, truth.max_column + 1):
            tv = truth.cell(r, c).value
            if tv in (None, "") or _norm(tv) == _norm(blank.cell(r, c).value):
                continue
            human[f"{__col(c)}{r}"] = tv
    gaps = sorted(set(human) - set(written))

    n = len(written)
    return {
        "doc": doc, "variant": variant, "sheet": sheet,
        "written": n, "match": match,
        "match_rate": (match / n) if n else 0.0,
        "human_cells": len(human),
        "covered": len(set(human) & set(written)),
        "coverage": (len(set(human) & set(written)) / len(human)) if human else 0.0,
        "diffs": diffs, "gaps": [(g, human[g]) for g in gaps],
    }


def __col(c: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(c)


def print_report(stats: dict[str, Any], max_rows: int = 20) -> None:
    print(f"\n■ {stats['doc']} / {stats['variant']} ({stats['sheet']})")
    print(f"  パイプライン充当セル: {stats['written']}")
    print(f"  公式様式と一致: {stats['match']}  → 一致率 {stats['match_rate']*100:.0f}%")
    print(f"  宅建士の案件固有セル: {stats['human_cells']} / "
          f"うち充当 {stats['covered']} → カバー率 {stats['coverage']*100:.0f}%")
    if stats["diffs"]:
        print(f"  不一致 {len(stats['diffs'])} 件:")
        for coord, pv, tv in stats["diffs"][:max_rows]:
            print(f"    {coord:6} pipe={str(pv)[:22]:24} truth={str(tv)[:26]}")
    if stats["gaps"]:
        print(f"  未カバー（手作業が残る）{len(stats['gaps'])} 件:")
        for coord, tv in stats["gaps"][:max_rows]:
            print(f"    {coord:6} truth={str(tv)[:34]}")


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    doc = "juyojiko"
    for a in sys.argv[1:]:
        if a.startswith("--doc"):
            doc = a.split("=", 1)[1] if "=" in a else sys.argv[sys.argv.index(a) + 1]
    if len(args) < 4:
        print(__doc__)
        return 1
    json_path, variant, blank_path, truth_path = args[:4]
    data = json.load(open(json_path, encoding="utf-8"))
    bc = (Juyojiko if doc == "juyojiko" else Keiyakusho).model_validate(data)
    print_report(compare(bc, variant, doc, blank_path, truth_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
