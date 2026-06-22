"""セルマップ整備用ツール（単一WB版）: 様式の印字ラベルから投入セル候補を提案する.

`wb_diff.py` は実例2通の差分で投入セルを特定するが、本ツールは **1通だけ**
（ブランク様式でも可）でラベルを手掛かりに投入セル候補を機械的に洗い出す。
借地説明書・全宅連様式など、まだ実例が揃わない新様式のマッピングを
「ゼロから手作業」ではなく「候補をレビューして確定」にするための補助。

重要: ここで出すのは **候補** であり、確定マッピングではない。実値での照合
（どのセルに地代・存続期間等が実際に入るか）を経てから cellmaps へ採用する。
「間違いなく」の原則上、未照合の候補をそのまま本番差込に使ってはならない。

使い方::
    python wb_probe.py 借地WB.xlsx "171-1.借地説明書" --preset shakuchi
    python wb_probe.py 全宅連WB.xlsx "重要事項説明書"   # プリセット無し=全ラベル走査
"""

from __future__ import annotations

import sys

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# 既知の様式ラベル（部分一致）。実例が無い段階でも当たりを付けられるよう用意。
PRESETS: dict[str, list[str]] = {
    "shakuchi": [
        "借地権の種類", "普通借地権", "定期借地権", "事業用定期借地権",
        "建物譲渡特約付", "借地権の登記", "存続期間", "地代", "地代の支払",
        "地代改定", "更新料", "借地権の譲渡", "譲渡の承諾", "増改築",
        "建物の築造", "底地", "地主", "土地所有者",
    ],
}

_SKIP = {"□", "■", "・", "（", "）", "年", "月", "日", "㎡", "円", "：", ":"}


def _is_label(v: object) -> bool:
    if not isinstance(v, str):
        return False
    s = v.strip()
    return len(s) >= 2 and s not in _SKIP


def _candidates(ws, r: int, c: int, max_row: int, max_col: int) -> list[str]:
    """ラベル右隣・下・近傍のチェック枠(□/■)を投入セル候補として返す。"""
    out: list[str] = []
    # 右方向に最初の空または記入対象セル
    for cc in range(c + 1, min(c + 8, max_col) + 1):
        v = ws.cell(r, cc).value
        if v in (None, ""):
            out.append(f"{get_column_letter(cc)}{r}")
            break
        if v in ("□", "■") or not _is_label(v):
            out.append(f"{get_column_letter(cc)}{r}")
            break
    # 直下セル
    if r + 1 <= max_row:
        bv = ws.cell(r + 1, c).value
        if bv in (None, "") or bv in ("□", "■"):
            out.append(f"{get_column_letter(c)}{r + 1}")
    # 近傍のチェック枠（同一行 ±6 列）
    for cc in range(max(1, c - 6), min(c + 6, max_col) + 1):
        if ws.cell(r, cc).value in ("□", "■"):
            co = f"{get_column_letter(cc)}{r}"
            if co not in out:
                out.append(co)
    return out


def probe(path: str, sheet: str, labels: list[str] | None) -> list[tuple]:
    ws = load_workbook(path, data_only=True)[sheet]
    mr, mc = ws.max_row, ws.max_column
    rows: list[tuple] = []
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if not _is_label(v):
                continue
            s = v.strip()
            if labels is not None and not any(key in s for key in labels):
                continue
            co = f"{get_column_letter(c)}{r}"
            cand = _candidates(ws, r, c, mr, mc)
            rows.append((co, s[:24], cand))
    return rows


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    preset = None
    for a in sys.argv[1:]:
        if a.startswith("--preset"):
            preset = a.split("=", 1)[1] if "=" in a else None
    if "--preset" in sys.argv:
        i = sys.argv.index("--preset")
        if i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith("--"):
            preset = sys.argv[i + 1]
            if preset in args:
                args.remove(preset)
    if len(args) < 2:
        print(__doc__)
        return 1
    path, sheet = args[0], args[1]
    labels = PRESETS.get(preset) if preset else None
    if preset and labels is None:
        print(f"未知のpreset: {preset}（候補: {', '.join(PRESETS)}）")
        return 1
    print(f"# {path} / {sheet} / preset={preset or '(全ラベル)'}")
    print("# ラベルセル, ラベル文言, 投入セル候補（要・実値照合で確定）")
    for co, label, cand in probe(path, sheet, labels):
        print(f"{co}\t{label}\t{cand}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
