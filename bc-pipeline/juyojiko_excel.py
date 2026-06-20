"""重要事項説明書を Excel で再現するレンダラ.

公式重説の章立て（取引態様 / 宅建業者・取引士 / 不動産の表示 /
法令制限 / 設備 / 区分管理 / 取引条件 / 容認事項 / 特約）を上から再現する。
用途地域・区域区分などの選択肢は ■/□ で表示する。
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from bc_schema import YOTO_OPTIONS, normalize_yoto
from juyojiko_schema import Juyojiko

KUIKI_OPTIONS = ["市街化区域", "市街化調整区域", "区域区分のされていない区域"]

_TITLE = Font(bold=True, size=16)
_SECTION = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="404040")
_LABEL = Font(bold=True, size=9)
_LABEL_FILL = PatternFill("solid", fgColor="EEEEEE")
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="center", wrap_text=True)
_VCENTER = Alignment(vertical="center")


class _Sheet:
    """行カーソルを持つ重説シート書き込みヘルパ."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        self.row = 1
        ws.column_dimensions["A"].width = 22   # ラベル
        ws.column_dimensions["B"].width = 60   # 値
        ws.column_dimensions["C"].width = 12

    def title(self, text: str) -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
        c = self.ws.cell(self.row, 1, text)
        c.font = _TITLE
        c.alignment = Alignment(horizontal="center")
        self.ws.row_dimensions[self.row].height = 28
        self.row += 2

    def section(self, text: str) -> None:
        self.ws.merge_cells(start_row=self.row, start_column=1, end_row=self.row, end_column=3)
        c = self.ws.cell(self.row, 1, text)
        c.font = _SECTION
        c.fill = _SECTION_FILL
        c.alignment = _VCENTER
        self.row += 1

    def kv(self, label: str, value: Any) -> None:
        """ラベル / 値の1行。値が None/空でも枠は出す（記入欄として）。"""
        lc = self.ws.cell(self.row, 1, label)
        lc.font = _LABEL
        lc.fill = _LABEL_FILL
        lc.border = _BORDER
        lc.alignment = _WRAP
        self.ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        vc = self.ws.cell(self.row, 2, _fmt(value))
        vc.border = _BORDER
        vc.alignment = _WRAP
        self.row += 1

    def checkbox(self, label: str, options: list[str], selected: str | None) -> None:
        """選択肢を ■/□ で1行に並べる（用途地域・区域区分等）。"""
        lc = self.ws.cell(self.row, 1, label)
        lc.font = _LABEL
        lc.fill = _LABEL_FILL
        lc.border = _BORDER
        lc.alignment = _WRAP
        parts = [f"{'■' if (selected and o == selected) else '□'} {o}" for o in options]
        self.ws.merge_cells(start_row=self.row, start_column=2, end_row=self.row, end_column=3)
        vc = self.ws.cell(self.row, 2, "　".join(parts))
        vc.border = _BORDER
        vc.alignment = _WRAP
        self.row += 1

    def listblock(self, label: str, items: list[str]) -> None:
        text = "\n".join(f"・{x}" for x in items) if items else ""
        self.kv(label, text)

    def gap(self) -> None:
        self.row += 1


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "有" if value else "無"
    if isinstance(value, (int, float)):
        return f"{value:,}"
    return str(value)


def _yen(value: Any) -> str:
    return f"{value:,} 円" if isinstance(value, (int, float)) else ""


def render(j: Juyojiko) -> bytes:
    """Juyojiko を Excel バイト列に描画する."""
    wb = Workbook()
    s = _Sheet(wb.active)
    wb.active.title = "重要事項説明書"

    bukken = j.bukken_type or (j.fudosan.bukken_type if j.fudosan else None) or ""
    s.title("重 要 事 項 説 明 書")
    s.kv("物件種別", bukken)
    s.kv("取引態様", j.torihiki_taiyo)
    s.gap()

    # 宅建業者・取引士（売主側）
    s.section("宅地建物取引業者（売主側）")
    g = j.gyosha
    s.kv("免許証番号", g.menkyo_no if g else None)
    s.kv("商号または名称", g.shomei if g else None)
    s.kv("主たる事務所の所在地", g.shozai if g else None)
    s.kv("代表者氏名 / 電話", f"{(g.daihyo if g else '') or ''}　{(g.tel if g else '') or ''}".strip())
    t = j.torikiishi
    s.kv("取引士 登録番号 / 氏名", f"{(t.toroku_no if t else '') or ''}　{(t.shimei if t else '') or ''}".strip())
    s.gap()

    # 当事者
    s.section("売主・買主")
    s.kv("売主（B）住所", j.urinushi.address if j.urinushi else None)
    s.kv("売主（B）氏名", j.urinushi.name if j.urinushi else None)
    s.kv("買主（C）住所", j.kainushi.address if j.kainushi else None)
    s.kv("買主（C）氏名", j.kainushi.name if j.kainushi else None)
    s.gap()

    # A 不動産の表示
    s.section("A 不動産の表示")
    f = j.fudosan
    if f:
        s.kv("住居表示", f.jukyo_hyoji)
        if (bukken == "区分") or f.senyuu or f.ittou_shozai:
            s.kv("一棟の建物 所在", f.ittou_shozai)
            s.kv("一棟の建物 構造", f.ittou_kozo)
            s.kv("一棟の建物 延床面積", f.ittou_enshoumenseki)
            se = f.senyuu
            s.kv("専有部分 家屋番号", se.kaoku_bango if se else None)
            s.kv("専有部分 建物の名称", se.meisho if se else None)
            s.kv("専有部分 種類 / 構造", f"{(se.shurui if se else '') or ''}　{(se.kozo if se else '') or ''}".strip())
            s.kv("専有部分 床面積", se.yukamenseki if se else None)
            s.kv("建築時期", se.chikujiki if se else None)
            for i, sk in enumerate(f.shikichiken or [], 1):
                s.kv(f"敷地権{i} 所在/地番", f"{sk.shozai or ''} {sk.chiban or ''}".strip())
                s.kv(f"敷地権{i} 地積/種類/割合",
                     f"{sk.chiseki or ''} / {sk.shikichiken_shurui or ''} / {sk.wariai or ''}")
        else:
            to = f.tochi
            s.kv("土地 所在・地番", to.shozai if to else None)
            s.kv("土地 地目", to.chimoku if to else None)
            s.kv("土地 地積（登記/実測）",
                 f"{(to.chiseki_toki if to else '') or ''} / {(to.chiseki_jissoku if to else '') or ''}")
            ta = f.tatemono
            s.kv("建物 家屋番号", ta.kaoku_bango if ta else None)
            s.kv("建物 種類 / 構造", f"{(ta.shurui if ta else '') or ''}　{(ta.kozo if ta else '') or ''}".strip())
            s.kv("建物 床面積", ta.yukamenseki if ta else None)
            s.kv("建築時期", ta.chikujiki if ta else None)
    s.kv("登記名義人（所有者）", j.touki_meigi)
    s.kv("第三者の占有（賃借人等）", j.senyuusha_uchi)
    s.gap()

    # Ⅰ-2 法令制限
    s.section("Ⅰ-2 都市計画法・建築基準法等に基づく制限")
    h = j.horei
    if h:
        s.kv("都市計画区域", h.toshikeikaku_kuiki)
        s.checkbox("区域区分", KUIKI_OPTIONS, h.kuiki_kubun)
        s.checkbox("用途地域", YOTO_OPTIONS, normalize_yoto(h.yoto))
        s.kv("建築基準法第22条区域", h.nijuni_jo)
        s.kv("防火/準防火", h.boka)
        s.kv("高度地区", h.kodo_chiku)
        s.listblock("その他の地域地区", h.chiiki_chiku)
        s.kv("指定建蔽率(%)", h.kenpei)
        s.kv("建蔽率の緩和等", h.kenpei_kanwa)
        s.kv("指定容積率(%)", h.yoseki)
        s.kv("前面道路による容積率制限", h.yoseki_zenmen_doro)
        s.kv("日影規制", h.nisshido)
        s.kv("接面道路", h.doro)
        s.listblock("その他の法令に基づく制限", h.other_horei)
    s.gap()

    # Ⅰ-4 設備
    s.section("Ⅰ-4 飲用水・電気・ガス・排水の整備状況")
    s.kv("整備状況（概要）", j.setsubi)
    s.gap()

    # Ⅰ-6 区分管理（区分のみ）
    if bukken == "区分" or j.kanri:
        s.section("Ⅰ-6 一棟の建物・敷地の管理（区分所有）")
        k = j.kanri
        if k:
            s.kv("通常の管理費（月額）", _yen(k.kanrihi_getsugaku))
            s.kv("修繕積立金（月額）", _yen(k.shuzen_getsugaku))
            s.kv("修繕積立金 積立累計", _yen(k.shuzen_tsumitate))
            s.kv("管理費 滞納額", _yen(k.kanrihi_taino))
            s.kv("修繕積立金 滞納額", _yen(k.shuzen_taino))
            s.kv("管理組合の名称", k.kanri_kumiai)
            s.kv("管理形態", k.kanri_keitai)
            s.kv("管理委託先", k.kanri_itakusaki)
            s.kv("専有部分の用途制限", k.yoto_seigen)
            s.kv("ペット飼育制限", k.pet_seigen)
        s.gap()

    # Ⅱ 取引条件
    s.section("Ⅱ 取引条件に関する事項")
    jo = j.joken
    if jo:
        s.kv("売買代金", _yen(jo.baibai_daikin))
        s.kv("うち消費税等相当額", _yen(jo.shohizei))
        s.kv("手付金", _yen(jo.tetsuke))
        s.kv("公租公課の清算起算日", jo.seisan_kisanbi)
        s.kv("違約金（売買代金の%）", jo.iyakukin_wariai)
        s.kv("担保責任/契約不適合の措置", jo.tanpo_sekinin)
        s.kv("融資利用の特約", jo.loan_tokuyaku)
    s.gap()

    # Ⅲ 容認事項
    s.section("Ⅲ その他重要な事項（容認事項）")
    s.listblock("容認事項", j.yonin_jiko)
    s.gap()

    # Ⅴ 特約
    s.section("Ⅴ 備考（特約事項）")
    s.listblock("特約事項", j.tokuyaku)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
