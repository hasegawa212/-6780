"""白紙 BC テンプレ（blank_36-1.xlsx / blank_37-1.xlsx）をスキーマから生成する.

実物テンプレが手に入ったら本スクリプトは不要。差し替えるだけでよい。
単体実行で両テンプレを出力する::

    python make_blank_templates.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from bc_schema import TEMPLATES, YOTO_OPTIONS, TemplateSpec

HERE = Path(__file__).resolve().parent

_TITLE_FONT = Font(bold=True, size=14)
_LABEL_FONT = Font(bold=True)
_LABEL_FILL = PatternFill("solid", fgColor="EEEEEE")
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_label(ws: Worksheet, coord: str, text: str) -> None:
    c = ws[coord]
    c.value = text
    c.font = _LABEL_FONT
    c.fill = _LABEL_FILL
    c.border = _BORDER
    c.alignment = Alignment(vertical="center")


def _style_value(ws: Worksheet, coord: str) -> None:
    c = ws[coord]
    c.border = _BORDER
    c.alignment = Alignment(vertical="center")


def build_template(spec: TemplateSpec, out_path: Path) -> Path:
    """1 テンプレ分の白紙 .xlsx を生成して保存する."""
    wb = Workbook()
    ws = wb.active
    ws.title = spec.sheet

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 44

    ws["B2"] = spec.title
    ws["B2"].font = _TITLE_FONT

    # 物件概要セル
    for cell in spec.cells:
        _style_label(ws, cell.label_cell, cell.label)
        _style_value(ws, cell.value_cell)

    # 用途地域チェックボックス見出し＋選択肢
    _style_label(ws, spec.yoto_header_cell,
                 f"用途地域 区分（{spec.yoto_checkbox_name}）")
    for i, opt in enumerate(YOTO_OPTIONS):
        row = spec.yoto_option_start_row + i
        ws[f"{spec.yoto_option_col}{row}"] = "□"
        ws[f"{spec.yoto_option_col}{row}"].alignment = Alignment(horizontal="center")
        ws[f"{spec.yoto_label_col}{row}"] = opt

    wb.save(out_path)
    return out_path


def build_all() -> list[Path]:
    out: list[Path] = []
    for spec in TEMPLATES.values():
        path = HERE / spec.template_file
        build_template(spec, path)
        out.append(path)
    return out


if __name__ == "__main__":
    for p in build_all():
        print(f"生成: {p.name}")
