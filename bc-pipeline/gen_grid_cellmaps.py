"""法令・地域地区チェック格子のセル座標を公式様式から生成するツール.

「(2) 地域・地区・街区等」(22件) と「(3) 都計法・建基法以外の法令」(61件) は
チェック枠が格子状に並ぶ。手書きで全座標を起こすと誤りやすいので、公式様式テンプレの
印字ラベル（"N.法令名"/"N.地域地区名"）左の□を走査し、`horei_master.normalize_horei`
で正規名に対応付けて **変種別の座標 dict を機械生成** する。

生成物（`cellmap_grids.py`）は静的リテラルとしてコミットし、実行時はWB非依存にする
（cellmaps「確証済みセルのみ」方針に合致）。様式が更新されたら本スクリプトを再実行して
`cellmap_grids.py` を再生成する。

使い方::
    python gen_grid_cellmaps.py 36-1=/path/36-1.xlsx 37-1=/path/37-1.xlsx 38-1=/path/38-1.xlsx \
        > cellmap_grids.py
"""

from __future__ import annotations

import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from horei_master import CHIIKI_CHIKU, OTHER_HOREI_LAWS, normalize_horei

SHEET = "重要事項説明書"
_NUM = re.compile(r"^\s*\d+\.")


def _clean(label: str) -> str:
    """先頭の "N." と末尾の括弧・空白を除いた法令/地区名を返す。"""
    return _NUM.sub("", label.strip()).rstrip("）) 　").strip()


def _left_box(ws, r: int, c: int) -> str | None:
    """ラベル左（最大4列）の最初の □/■ 枠の座標を返す。"""
    for cc in range(c - 1, max(0, c - 5), -1):
        if ws.cell(r, cc).value in ("□", "■"):
            return f"{get_column_letter(cc)}{r}"
    return None


def _grid(path: str, master: list[str]) -> dict[str, str]:
    """様式から {正規名: チェック枠座標} を master 記載順で返す。"""
    ws = load_workbook(path, data_only=True)[SHEET]
    out: dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not (isinstance(v, str) and _NUM.match(v.strip())):
                continue
            name = normalize_horei(_clean(v))
            if name in master and name not in out:
                box = _left_box(ws, r, c)
                if box:
                    out[name] = box
    # master 順に並べ替えて返す
    return {k: out[k] for k in master if k in out}


def _emit(var_name: str, by_variant: dict[str, dict[str, str]]) -> str:
    lines = [f"{var_name} = {{"]
    for variant, mapping in by_variant.items():
        lines.append(f'    "{variant}": {{')
        for name, coord in mapping.items():
            lines.append(f'        "{name}": "{coord}",')
        lines.append("    },")
    lines.append("}")
    return "\n".join(lines)


def main() -> int:
    pairs = dict(a.split("=", 1) for a in sys.argv[1:] if "=" in a)
    if not pairs:
        print(__doc__)
        return 1
    chiiki = {v: _grid(p, CHIIKI_CHIKU) for v, p in pairs.items()}
    horei = {v: _grid(p, OTHER_HOREI_LAWS) for v, p in pairs.items()}
    # 件数を stderr に出して取りこぼしを検知（36-1/38-1=61, 37-1=60, 地区=22 が期待値）
    for v in pairs:
        print(f"# {v}: 地域地区 {len(chiiki[v])}/{len(CHIIKI_CHIKU)}, "
              f"法令 {len(horei[v])}/{len(OTHER_HOREI_LAWS)}", file=sys.stderr)
    print('"""法令・地域地区チェック格子のセル座標（変種別）.')
    print()
    print("gen_grid_cellmaps.py が公式様式テンプレから自動生成。手で編集しない。")
    print("様式更新時は gen_grid_cellmaps.py を再実行して本ファイルを差し替える。")
    print('"""')
    print()
    print("from __future__ import annotations")
    print()
    print("# (2) 地域・地区・街区等（horei_master.CHIIKI_CHIKU 22件）")
    print(_emit("CHIIKI_CHIKU_MARKS", chiiki))
    print()
    print("# (3) 都計法・建基法以外の法令（horei_master.OTHER_HOREI_LAWS 61件）")
    print("# 37-1 実様式は最新の「生物多様性増進法」を欠く旧版（60件）。")
    print("# 38-1 は同法を挿入する分、以降の行が 37-1 と相違する。")
    print(_emit("OTHER_HOREI_MARKS", horei))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
