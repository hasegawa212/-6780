"""BC 自動生成サービス（FastAPI）.

AB 側（仕入れ）重要事項説明書を読み込み、BC 側（B→C 転売）の重説を生成する。

エンドポイント:
  GET  /health   … 死活監視
  POST /extract  … AB重説(PDF/画像/テキスト) → 構造化 JSON（Juyojiko）
  POST /generate … AB重説JSON ＋案件マスタ → BC重説(.xlsx) を base64 で返す

環境変数:
  ANTHROPIC_API_KEY  … Claude（/extract のみ必須）
  ANTHROPIC_BASE_URL … 社内 LiteLLM プロキシ等に向ける場合（任意）
  CLAUDE_MODEL       … 既定 claude-opus-4-8
  CLAUDE_MAX_TOKENS  … 既定 4000
"""

from __future__ import annotations

import base64
import io
import json
import os
from typing import Any

from pathlib import Path

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

import approval
import auth
import bundle
import cellmaps
import validate
import juyojiko_excel
import keiyaku_excel
import wb_fill
from bc_schema import YOTO_OPTIONS, normalize_yoto, resolve_bukken
from bc_transform import transform_ab_to_bc, transform_keiyaku_ab_to_bc
from juyojiko_schema import (
    FudosanHyoji,
    HoreiSeigen,
    Juyojiko,
    TatemonoHyoji,
    TochiHyoji,
    TorihikiJoken,
)
from keiyaku_schema import Keiyakusho

MODEL = os.environ.get("CLAUDE_MODEL", "claude-opus-4-8")
MAX_TOKENS = int(os.environ.get("CLAUDE_MAX_TOKENS", "4000"))

app = FastAPI(title="BC自動生成サービス", version="0.2.0")

_WEBUI = Path(__file__).parent / "webui" / "index.html"
_LOGIN = Path(__file__).parent / "webui" / "login.html"

# 認証なしでアクセスできるパス（ログイン画面・死活監視・favicon）。
_PUBLIC_PATHS = {"/login", "/logout", "/health", "/favicon.ico"}


@app.middleware("http")
async def _auth_gate(request: Request, call_next):  # type: ignore[no-untyped-def]
    """認証有効時、未ログインのアクセスを遮断する。ブラウザ操作(GET)はログイン画面へ
    リダイレクト、API(POST等)は 401 JSON を返す。認証無効時は素通り（後方互換）。"""
    if not auth.is_enabled() or request.url.path in _PUBLIC_PATHS:
        return await call_next(request)
    if auth.current_user(request.cookies):
        return await call_next(request)
    accepts_html = "text/html" in request.headers.get("accept", "")
    if request.method == "GET" and accepts_html:
        nxt = request.url.path
        return RedirectResponse(f"/login?next={nxt}", status_code=303)
    return JSONResponse({"detail": "ログインが必要です。"}, status_code=401)


def _set_session_cookie(resp: Any, username: str) -> None:
    resp.set_cookie(
        auth.COOKIE_NAME, auth.create_session(username),
        httponly=True, samesite="lax", secure=auth.is_secure_cookie(), path="/",
    )


@app.get("/", response_class=HTMLResponse)
def webui() -> str:
    """ブラウザ用の操作画面（AB読取→BC情報入力→BC一式ダウンロード）。"""
    if _WEBUI.exists():
        return _WEBUI.read_text(encoding="utf-8")
    return "<h1>BC自動生成サービス</h1><p>webui/index.html が見つかりません。</p>"


@app.get("/login", response_class=HTMLResponse)
def login_page(next: str = "/", error: str = "") -> HTMLResponse:
    """ログイン画面。認証が無効なら操作画面へ戻す。"""
    if not auth.is_enabled():
        return RedirectResponse("/", status_code=303)  # type: ignore[return-value]
    html = _LOGIN.read_text(encoding="utf-8") if _LOGIN.exists() else \
        "<form method=post action=/login>ID<input name=username> " \
        "PW<input name=password type=password><button>ログイン</button></form>"
    banner = '<div class="err">IDまたはパスワードが違います。</div>' if error else ""
    html = html.replace("<!--ERROR-->", banner).replace("__NEXT__", next or "/")
    return HTMLResponse(html)


@app.post("/login")
async def login_submit(request: Request) -> Any:
    """ログイン処理。成功でセッションクッキーを発行し操作画面へ。

    フォームは application/x-www-form-urlencoded。request.form() で読むため
    ファイルアップロード用の python-multipart には依存しない。
    """
    from urllib.parse import parse_qs
    body = (await request.body()).decode("utf-8")
    form = parse_qs(body, keep_blank_values=True)
    username = (form.get("username") or [""])[0]
    password = (form.get("password") or [""])[0]
    nxt = (form.get("next") or ["/"])[0]
    if not auth.authenticate(username, password):
        return RedirectResponse("/login?error=1", status_code=303)
    dest = nxt if nxt.startswith("/") else "/"
    resp = RedirectResponse(dest, status_code=303)
    _set_session_cookie(resp, username)
    return resp


@app.post("/logout")
@app.get("/logout")
def logout() -> Any:
    """ログアウト（セッションクッキーを破棄）。"""
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie(auth.COOKIE_NAME, path="/")
    return resp


@app.get("/me")
def me(request: Request) -> dict[str, Any]:
    """ログイン中ユーザー情報（webui のユーザー表示用）。"""
    if not auth.is_enabled():
        return {"auth_enabled": False, "username": None, "display_name": None}
    user = auth.current_user(request.cookies)
    return {
        "auth_enabled": True,
        "username": user,
        "display_name": auth.display_name(user) if user else None,
    }


# ── リクエスト/レスポンス ─────────────────────────────────────
class GenerateReq(BaseModel):
    model_config = ConfigDict(extra="allow")

    doc_type: str = "juyojiko"          # juyojiko（重説） / keiyaku（契約書） / package（両方）
    # 本番ワークブック差込: テンプレ変種（36-1 / 37-1 / 38-1）。指定時は差込を試みる。
    template: str | None = None
    template_base64: str | None = None  # 御社ワークブックを直接渡す場合
    # 新方式: AB 書類の構造化 JSON（/extract の出力）
    ab: dict[str, Any] | None = None        # 重説 JSON（package では重説シート用）
    ab_keiyaku: dict[str, Any] | None = None  # 契約書 JSON（package で契約書シート用）
    # 旧方式（手順書 curl 互換）: 最小フィールド（重説のみ）
    bukken: str | None = None
    extracted: dict[str, Any] | None = None
    # 案件マスタ（BC 側の当事者・代金など）
    deal_master: dict[str, Any] = {}
    filename: str | None = None


class GenerateResp(BaseModel):
    filename: str
    bukken: str
    xlsx_base64: str
    # 発行前チェック（記入漏れ・不整合・御社標準からの逸脱）。空なら問題なし。
    warnings: list[dict[str, str]] = []


class ExtractReq(BaseModel):
    doc_type: str = "juyojiko"          # juyojiko（重説） / keiyaku（売買契約書）
    bukken: str | None = None
    text: str | None = None
    file_base64: str | None = None
    mime: str = "application/pdf"


class ExtractResp(BaseModel):
    extracted: dict[str, Any]
    # 自動読取が部分的/不能だった場合の非致命メッセージ（空なら全て正常）。
    # UI はこれを注意表示しつつ、手入力で先へ進める（読取失敗でも止めない）。
    warning: str = ""


# ── /health ───────────────────────────────────────────────────
@app.get("/health")
def health() -> dict[str, Any]:
    tdir = os.environ.get("BC_TEMPLATE_DIR", "templates")
    templates = sorted(
        p.stem for p in __import__("pathlib").Path(tdir).glob("*.xlsx")
    ) if os.path.isdir(tdir) else []
    return {
        "status": "ok",
        "version": "0.2.0",
        "model": MODEL,
        "bukken": ["戸建", "区分"],
        # 設定の見える化（秘密情報は出さない）
        "api_key_configured": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "base_url": os.environ.get("ANTHROPIC_BASE_URL", "https://api.anthropic.com"),
        "template_dir": tdir,
        "templates_available": templates,   # 例: ["36-1","37-1","38-1"]
    }


# ── /reference（法令制限の正式名称マスタ）──────────────────────
@app.get("/reference")
def reference() -> dict[str, Any]:
    import horei_master

    return {
        "yoto": horei_master.YOTO_OPTIONS,
        "chiiki_chiku": horei_master.CHIIKI_CHIKU,
        "other_horei": horei_master.OTHER_HOREI_LAWS,
    }


@app.get("/masters")
def masters() -> dict[str, Any]:
    """アプリのプリセット用マスタ（売主業者B＝御社・御社取引士・媒介業者）。

    アプリはこれを使って「選ぶだけ」のドロップダウンを作り、入力を買主C・価格に絞る。
    """
    import house_style

    return {
        "seller_b": house_style.SELLER_B_MASTER,
        "seller_b_torikiishi": house_style.SELLER_B_TORIKIISHI,
        "baikai_gyosha": house_style.BAIKAI_GYOSHA_MASTER,
    }


# ── /bundle（添付書類のPDF結合）────────────────────────────────
class BundleReq(BaseModel):
    attachments: list[str]        # base64 PDF（結合する順）
    filename: str | None = None


class BundleResp(BaseModel):
    filename: str
    page_count: int
    pdf_base64: str


@app.post("/bundle", response_model=BundleResp)
def bundle_pdfs(req: BundleReq) -> BundleResp:
    if not req.attachments:
        raise HTTPException(status_code=400, detail="attachments が空です。")
    try:
        pdfs = [base64.b64decode(a) for a in req.attachments]
        merged, pages = bundle.merge_pdfs(pdfs)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"PDF結合に失敗: {e}") from e
    return BundleResp(
        filename=req.filename or "添付書類束.pdf",
        page_count=pages,
        pdf_base64=base64.b64encode(merged).decode("ascii"),
    )


# ── /approval（Slack承認 ✅/❌ の判定）─────────────────────────
class ApprovalReq(BaseModel):
    model_config = ConfigDict(extra="allow")

    # Slack の url_verification ハンドシェイク用
    type: str | None = None
    challenge: str | None = None
    # 簡易形式: {"reaction":"✅"}。Events API の場合は event.reaction を読む。
    reaction: str | None = None
    event: dict[str, Any] | None = None


@app.post("/approval")
def approval_hook(req: ApprovalReq) -> dict[str, Any]:
    # Slack Events API の URL 検証（challenge をそのまま返す）
    if req.type == "url_verification" and req.challenge:
        return {"challenge": req.challenge}
    reaction = approval.reaction_from_payload(req.model_dump(exclude_none=True))
    decision = approval.decide(reaction)
    return {"decision": decision, "approved": decision == "approve", "reaction": reaction}


# ── /generate ─────────────────────────────────────────────────
def _legacy_to_juyojiko(bukken: str, extracted: dict[str, Any]) -> Juyojiko:
    """手順書 curl の最小フィールドを Juyojiko へマップ（後方互換）."""
    key = resolve_bukken(bukken)
    shozai = extracted.get("shozai")
    return Juyojiko(
        bukken_type=key,
        fudosan=FudosanHyoji(
            bukken_type=key,
            jukyo_hyoji=shozai,
            tochi=TochiHyoji(shozai=shozai) if key == "戸建" else None,
            tatemono=TatemonoHyoji() if key == "戸建" else None,
            senyuu=TatemonoHyoji() if key == "区分" else None,
            ittou_shozai=shozai if key == "区分" else None,
        ),
        horei=HoreiSeigen(
            kuiki_kubun=extracted.get("kuiki"),
            yoto=normalize_yoto(extracted.get("yoto")),
            nijuni_jo=extracted.get("nijuni_jo"),
            kenpei=extracted.get("kenpei"),
            yoseki=extracted.get("yoseki"),
        ),
        joken=TorihikiJoken(),
    )


def _shozai_of(f: Any) -> str:
    if not f:
        return "物件"
    tochi = getattr(f, "tochi", None)
    return (getattr(f, "jukyo_hyoji", None) or getattr(f, "ittou_shozai", None)
            or (getattr(tochi, "shozai", None) if tochi else None) or "物件")


def _filename(prefix: str, bukken: str, f: Any, override: str | None) -> str:
    if override:
        return override
    safe = "".join(c for c in str(_shozai_of(f)) if c not in r'\/:*?"<>|').strip()
    return f"{prefix}_{bukken}_{safe[:40]}.xlsx"


def _generate_juyojiko(req: GenerateReq) -> GenerateResp:
    if req.ab is not None:
        try:
            ab = Juyojiko.model_validate(req.ab)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"ab の解析に失敗: {e}") from e
    elif req.extracted is not None:
        if not req.bukken:
            raise HTTPException(status_code=400, detail="bukken が必要です。")
        try:
            ab = _legacy_to_juyojiko(req.bukken, req.extracted)
        except KeyError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
    else:
        raise HTTPException(status_code=400, detail="ab または extracted が必要です。")

    bc = transform_ab_to_bc(ab, req.deal_master)
    bukken = bc.bukken_type or (bc.fudosan.bukken_type if bc.fudosan else None) or "区分"

    # 本番ワークブックがあれば差込（最も忠実）。無ければ自作 Excel にフォールバック。
    template = _try_template_bytes(req)
    variant = _resolve_variant(req, template)  # 様式は明示指定 or A1から自動判定
    edition = _kubun_edition(template, variant)  # 区分B版テンプレはB版座標へ写像
    if template is not None and variant in cellmaps.JUYOJIKO_BUILDERS:
        try:
            sv, sc = cellmaps.build_juyojiko(variant, bc, edition=edition)
            av, ac = cellmaps.build_aux(bc)
            xlsx, _ = wb_fill.fill_workbook(template, {**sv, **av}, {**sc, **ac})
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"ワークブック差込に失敗: {e}") from e
        prefix = f"BC重説_{variant}"
    else:
        try:
            xlsx = juyojiko_excel.render(bc)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"重説生成に失敗: {e}") from e
        prefix = "BC重説"

    return GenerateResp(
        filename=_filename(prefix, bukken, bc.fudosan, req.filename),
        bukken=bukken,
        xlsx_base64=base64.b64encode(xlsx).decode("ascii"),
        warnings=validate.validate_juyojiko(bc),
    )


def _kubun_edition(template: bytes | None, variant: str | None) -> str:
    """区分(37-1/38-1)テンプレの様式版 'A'/'B' を判定する。区分以外・判定不能は 'A'。"""
    if template is None or variant not in ("37-1", "38-1"):
        return "A"
    try:
        wb = load_workbook(io.BytesIO(template), data_only=True)
        ws = wb[cellmaps.JUYOJIKO_SHEET] if cellmaps.JUYOJIKO_SHEET in wb.sheetnames else None
    except Exception:  # noqa: BLE001
        return "A"
    ed = cellmaps.detect_kubun_edition(ws) if ws is not None else "A"
    return "B" if ed == "B" else "A"


def _resolve_variant(req: GenerateReq, template: bytes | None) -> str | None:
    """様式を決める: 明示指定（req.template）を優先、無ければWBのA1から自動判定。"""
    if req.template:
        return req.template
    if template is not None:
        return wb_fill.detect_variant(template)
    return None


def _try_template_bytes(req: GenerateReq) -> bytes | None:
    """本番ワークブックのテンプレ実体を返す（無ければ None）。"""
    if req.template_base64:
        return base64.b64decode(req.template_base64)
    if req.template:
        tdir = os.environ.get("BC_TEMPLATE_DIR", "templates")
        return wb_fill.load_template(tdir, req.template)
    return None


def _generate_keiyaku(req: GenerateReq) -> GenerateResp:
    if req.ab is None:
        raise HTTPException(status_code=400, detail="契約書には ab（契約書JSON）が必要です。")
    try:
        ab = Keiyakusho.model_validate(req.ab)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"ab の解析に失敗: {e}") from e
    bc = transform_keiyaku_ab_to_bc(ab, req.deal_master)
    bukken = bc.bukken_type or (bc.fudosan.bukken_type if bc.fudosan else None) or "戸建"

    # 本番ワークブックがあれば差込（最も忠実）。無ければ自作 Excel にフォールバック。
    template = _try_template_bytes(req)
    variant = _resolve_variant(req, template)  # 様式は明示指定 or A1から自動判定
    if template is not None and variant in cellmaps.KEIYAKU_BUILDERS:
        try:
            sv, sc = cellmaps.build_keiyaku(variant, bc)
            av, ac = cellmaps.build_aux(bc)
            xlsx, _ = wb_fill.fill_workbook(template, {**sv, **av}, {**sc, **ac})
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"ワークブック差込に失敗: {e}") from e
        prefix = f"BC契約書_{variant}"
    else:
        try:
            xlsx = keiyaku_excel.render(bc)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"契約書生成に失敗: {e}") from e
        prefix = "BC契約書"

    return GenerateResp(
        filename=_filename(prefix, bukken, bc.fudosan, req.filename),
        bukken=bukken,
        xlsx_base64=base64.b64encode(xlsx).decode("ascii"),
        warnings=validate.validate_keiyaku(bc),
    )


def _generate_package(req: GenerateReq) -> GenerateResp:
    """重説シートと契約書シートを1つの本番ワークブックへ同時差込する。

    req.ab=重説JSON、req.ab_keiyaku=契約書JSON、req.template=変種、案件マスタを共用。
    本番ワークブック（両シートを含む）が必須。
    """
    template = _try_template_bytes(req)
    variant = _resolve_variant(req, template)  # 様式は明示指定 or A1から自動判定
    if template is None or variant not in cellmaps.JUYOJIKO_BUILDERS:
        raise HTTPException(
            status_code=400,
            detail="package には本番ワークブック（A1様式が36-1/37-1/38-1）が必要です。")
    # 契約書シートはA/B版で同一レイアウト（実WBで確認）。重説のみB版座標へ写像する。
    edition = _kubun_edition(template, variant)
    if req.ab is None or req.ab_keiyaku is None:
        raise HTTPException(
            status_code=400, detail="package には ab（重説）と ab_keiyaku（契約書）が必要です。")
    try:
        bc_j = transform_ab_to_bc(Juyojiko.model_validate(req.ab), req.deal_master)
        bc_k = transform_keiyaku_ab_to_bc(
            Keiyakusho.model_validate(req.ab_keiyaku), req.deal_master)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"ab の解析に失敗: {e}") from e

    sv_j, sc_j = cellmaps.build_juyojiko(variant, bc_j, edition=edition)
    sv_k, sc_k = cellmaps.build_keiyaku(variant, bc_k)
    av, ac = cellmaps.build_aux(bc_j)
    sheet_values = {**sv_j, **sv_k, **av}      # 重説 + 契約書 + 補助シート
    sheet_clear = {**sc_j, **sc_k, **ac}
    try:
        xlsx, _ = wb_fill.fill_workbook(template, sheet_values, sheet_clear)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"ワークブック差込に失敗: {e}") from e

    bukken = bc_j.bukken_type or (bc_j.fudosan.bukken_type if bc_j.fudosan else None) or "区分"
    return GenerateResp(
        filename=_filename(f"BC一式_{variant}", bukken, bc_j.fudosan, req.filename),
        bukken=bukken,
        xlsx_base64=base64.b64encode(xlsx).decode("ascii"),
        warnings=validate.validate_bc(juyojiko=bc_j, keiyaku=bc_k),
    )


@app.post("/generate", response_model=GenerateResp)
def generate(req: GenerateReq) -> GenerateResp:
    if req.doc_type == "package":
        return _generate_package(req)
    if req.doc_type == "keiyaku":
        return _generate_keiyaku(req)
    if req.doc_type == "juyojiko":
        return _generate_juyojiko(req)
    raise HTTPException(
        status_code=400,
        detail=f"未知の doc_type: {req.doc_type}（juyojiko / keiyaku / package）")


# ── /extract ──────────────────────────────────────────────────
_EXTRACT_SYS = (
    "あなたは日本の不動産の重要事項説明書（35条書面）を構造化する抽出エンジンです。"
    "与えられた重説（PDF/画像/テキスト）から、次の JSON 構造で読み取れる項目を返してください。"
    "読み取れない項目は null、配列は空配列に。推測で埋めないこと。前置き不要、JSON のみ。\n\n"
    "{\n"
    '  "bukken_type": "戸建|区分",\n'
    '  "torihiki_taiyo": "取引態様（例: 売買・媒介）",\n'
    '  "gyosha": {"menkyo_no":..,"menkyo_date":..,"shozai":..,"tel":..,"shomei":..,"daihyo":..},\n'
    '  "torikiishi": {"toroku_no":..,"shimei":..,"jimusho":..,"jimusho_shozai":..,"tel":..},\n'
    '  "urinushi": {"address":..,"name":..,"biko":..},\n'
    '  "fudosan": {"bukken_type":..,"jukyo_hyoji":..,"fuzoku_tatemono":"附属建物の有無(有/無)",'
    '"fuzoku_tatemono_detail":"附属建物の詳細",'
    '"tochi":{"shozai":..,"chimoku":..,"chiseki_toki":..,"chiseki_jissoku":..},'
    '"tatemono":{"kaoku_bango":..,"shurui":..,"kozo":..,"yukamenseki":..,"chikujiki":..},'
    '"ittou_shozai":..,"ittou_kozo":..,"ittou_enshoumenseki":..,'
    '"senyuu":{"kaoku_bango":..,"meisho":..,"shurui":..,"kozo":..,"yukamenseki":..,"chikujiki":..},'
    '"shikichiken":[{"shozai":..,"chiban":..,"chiseki":..,"shikichiken_shurui":..,"wariai":..}]},\n'
    '  "touki_meigi": "登記名義人（所有者）",\n'
    '  "senyuusha_uchi": "第三者占有(賃借人)の有無・概要",\n'
    '  "horei": {"toshikeikaku_kuiki":"都市計画区域内|外",'
    '"kuiki_kubun":"市街化区域|市街化調整区域|区域区分のされていない区域",'
    f'"yoto":"用途地域(次のいずれか: {", ".join(YOTO_OPTIONS)})",'
    '"nijuni_jo":true/false,"boka":..,"kodo_chiku":..,"chiiki_chiku":[..],'
    '"kenpei":整数,"kenpei_kanwa":..,"yoseki":整数,"yoseki_zenmen_doro":..,'
    '"nisshido":..,"doro":"接面道路の概要(備考)",'
    '"doro_hoko":"接面道路の方向(南/南西等)","doro_haba":"幅員","doro_setsudo":"接道の長さ",'
    '"shikichi_saitei":"敷地面積の最低限度","suigai_shozai":"水害ハザード所在地の説明",'
    '"other_horei":[..]},\n'
    '  "setsubi": "飲用水・電気・ガス・排水の概要",\n'
    '  "setsubi_detail": {"suidou":"公営水道|私営水道|井戸",'
    '"gas":"都市ガス|個別プロパン|集中プロパン",'
    '"osui":"公共下水|個別浄化槽|集中浄化槽|汲取式",'
    '"zassui":"公共下水|個別浄化槽|集中浄化槽|側溝等|浸透式",'
    '"denryoku":"電力会社名","biko":"設備の備考"},\n'
    '  "saigai": {"zosei_bosai":true/false,"dosha_keikai":true/false,'
    '"dosha_tokubetsu":true/false,"tsunami_keikai":true/false,"tsunami_tokubetsu":true/false,'
    '"taishin_shindan":"耐震診断 有=true","sekimen_kiroku":"石綿調査記録 有=true",'
    '"kozui":"水害洪水ハザード 有=true","naisui":"内水ハザード 有=true","takashio":"高潮 有=true"},\n'
    '  "kakunin": {"kenchiku_bango":"建築確認番号","kenchiku_date":"建築確認交付年月日",'
    '"kensa_bango":"検査済証番号","kensa_date":"検査済証交付年月日"},\n'
    '  "touki": {"tochi_shoyusha_jusho":"土地所有者住所","tochi_shoyusha_shimei":"土地所有者氏名",'
    '"tochi_otsuku":"土地乙区","tatemono_shoyusha_jusho":"建物所有者住所",'
    '"tatemono_shoyusha_shimei":"建物所有者氏名","tatemono_otsuku":"建物乙区"},\n'
    '  "kanri": {"kanrihi_getsugaku":整数,"shuzen_getsugaku":整数,"shuzen_tsumitate":整数,'
    '"kanrihi_taino":整数,"shuzen_taino":整数,"kanri_kumiai":..,"kanri_keitai":..,'
    '"kanri_itakusaki":..,"yoto_seigen":..,"pet_seigen":..},\n'
    '  "shakuchi": {"shakuchiken_shurui":"普通/一般定期/事業用定期/建物譲渡特約付",'
    '"toki_umu":"有/無","sonzoku_kikan":..,"keiyaku_shiki":..,"keiyaku_manryo":..,'
    '"jidai_kingaku":整数,"jidai_tani":"月額/年額","jidai_shiharai":..,"jidai_kaitei":..,'
    '"koshin_ryo":..,"joto_shodaku":..,"kenchiku_seigen":..,'
    '"teichi_shoyusha_jusho":..,"teichi_shoyusha_shimei":..,"biko":..},'
    "  # ↑借地権付き建物のときのみ。所有権物件では省略可\n"
    '  "joken": {"baibai_daikin":整数,"tochi_kakaku":"うち土地価格(整数)",'
    '"tatemono_kakaku":"うち建物価格(整数)","shohizei":整数,"tetsuke":整数,'
    '"seisan_kisanbi":..,"iyakukin_wariai":整数,"tanpo_sekinin":..,"loan_tokuyaku":true/false},\n'
    '  "seisan_biko": "公租公課の清算に関する備考",\n'
    '  "yonin_jiko": ["容認事項..."],\n'
    '  "tokuyaku": ["特約事項..."]\n'
    "}\n"
    "金額は円の整数（カンマ無し）。建蔽率・容積率は%の整数。"
)


_EXTRACT_SYS_KEIYAKU = (
    "あなたは日本の不動産売買契約書（FRK標準書式）を構造化する抽出エンジンです。"
    "与えられた契約書（PDF/画像/テキスト）から、次の JSON 構造で読み取れる項目を返してください。"
    "読み取れない項目は null、配列は空配列に。推測で埋めないこと。前置き不要、JSON のみ。\n\n"
    "{\n"
    '  "bukken_type": "戸建|区分",\n'
    '  "urinushi": {"address":..,"name":..},\n'
    '  "kainushi": {"address":..,"name":..},\n'
    '  "gyosha": {"shomei":..,"shozai":..,"tel":..,"daihyo":..},\n'
    '  "torikiishi": {"shimei":..,"toroku_no":..},\n'
    '  "fudosan": {"bukken_type":..,"jukyo_hyoji":..,'
    '"tochi":{"shozai":..,"chimoku":..,"chiseki_toki":..,"chiseki_jissoku":..},'
    '"tatemono":{"kaoku_bango":..,"shurui":..,"kozo":..,"yukamenseki":..,"chikujiki":..},'
    '"ittou_shozai":..,"senyuu":{"kaoku_bango":..,"yukamenseki":..},'
    '"shikichiken":[{"shozai":..,"chiban":..,"chiseki":..,"wariai":..}]},\n'
    '  "daikin": {"baibai_daikin":整数,"shohizei":整数,"tetsuke":整数,'
    '"uchikin1":整数,"uchikin1_date":..,"uchikin2":整数,"uchikin2_date":..,'
    '"zankin":整数,"zankin_date":..,"iyakukin_wariai":"違約金の額(売買代金の%。整数)"},\n'
    '  "hikiwatashi_date": "引渡し日","seisan_kisanbi":"公租公課の清算起算日",'
    '"keiyaku_date":"契約締結日",\n'
    '  "loan_tokuyaku": true/false,"loan_kingaku":整数,"loan_shonin_date":..,'
    '"loan_kaijo_date":"融資特約に基づく契約解除期日",\n'
    '  "tokuyaku": ["特約事項..."],\n'
    '  "jokan": [{"jo":"第1条","midashi":"見出し","honbun":"本文"}]\n'
    "}\n"
    "金額は円の整数（カンマ無し）。約款(jokan)は条ごとに分けて、本文も読み取れる範囲で含める。"
)


# 直接PDF送信の上限（Anthropic の 32MB/100頁 とリクエスト全体上限に対し安全側）。
# 超えたら本文テキスト送信 or 頁分割にフォールバックする。
_MAX_DIRECT_BYTES = 18 * 1024 * 1024
_MAX_DIRECT_PAGES = 95
_BATCH_MAX_BYTES = 15 * 1024 * 1024


def _instruction(doc_type: str) -> dict[str, Any]:
    doc = "売買契約書" if doc_type == "keiyaku" else "重要事項説明書"
    return {"type": "text", "text": f"次の{doc}から項目を抽出してください。"}


def _pdf_stats(b64: str) -> tuple[bytes, int, int | None, str]:
    """PDFの (バイト列, サイズ, 頁数, 抽出テキスト) を返す。pypdf 不在/破損時は頁数None・空文字。"""
    raw = base64.b64decode(b64)
    pages: int | None = None
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        pages = len(reader.pages)
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
    except BaseException:  # noqa: BLE001 破損PDF・暗号化・pypdf不備(pyo3 panic等)でも落とさない
        pass
    return raw, len(raw), pages, text


def _slice_pdf_b64(raw: bytes, start: int, end: int) -> str | None:
    """PDFの [start, end) 頁だけの小PDFをbase64で返す。失敗時 None。"""
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(io.BytesIO(raw))
        writer = PdfWriter()
        for i in range(start, min(end, len(reader.pages))):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except BaseException:  # noqa: BLE001 pypdf不備(pyo3 panic等)でも落とさない
        return None


def _call_claude_json(doc_type: str, pieces: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Claudeを1回呼び、JSONを返す。API失敗/JSON崩れは None（呼び出し側で継続）。"""
    from anthropic import Anthropic
    system = _EXTRACT_SYS_KEIYAKU if doc_type == "keiyaku" else _EXTRACT_SYS
    client = Anthropic(max_retries=4, timeout=180.0)
    msg = client.messages.create(
        model=MODEL, max_tokens=MAX_TOKENS, system=system,
        messages=[{"role": "user", "content": [_instruction(doc_type), *pieces]}],
    )
    raw = "".join(b.text for b in msg.content
                  if getattr(b, "type", None) == "text").strip()
    return _parse_json_loose(raw)


def _parse_json_loose(raw: str) -> dict[str, Any] | None:
    """コードフェンス除去＋最初の {...} 抽出まで試すゆるいJSON解析。無理なら None。"""
    if not raw:
        return None
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        raw = raw[4:] if raw.lstrip().startswith("json") else raw
    raw = raw.strip()
    for candidate in (raw, raw[raw.find("{"): raw.rfind("}") + 1] if "{" in raw else ""):
        try:
            obj = json.loads(candidate)
            return obj if isinstance(obj, dict) else None
        except (json.JSONDecodeError, ValueError):
            continue
    return None


def _merge_extracted(a: dict[str, Any], b: dict[str, Any]) -> dict[str, Any]:
    """2つの抽出結果を統合。スカラは先勝ち（先の非空を保持）、リストは連結重複除去、辞書は再帰。"""
    out = dict(a)
    for k, v in b.items():
        if k not in out or out[k] in (None, "", [], {}):
            out[k] = v
        elif isinstance(out[k], dict) and isinstance(v, dict):
            out[k] = _merge_extracted(out[k], v)
        elif isinstance(out[k], list) and isinstance(v, list):
            seen = {json.dumps(x, ensure_ascii=False, sort_keys=True) for x in out[k]}
            out[k] += [x for x in v
                       if json.dumps(x, ensure_ascii=False, sort_keys=True) not in seen]
    return out


def _extract_robust(req: ExtractReq) -> tuple[dict[str, Any], str]:
    """(抽出結果, 警告) を返す。どんな入力でも例外を投げない。
    重い/大きいPDFは 直接送信→本文テキスト→頁分割 の順に自動フォールバックし、
    一部が失敗しても取れた分を返す。全滅時は空dict＋理由（手入力で続行可能）。"""
    try:
        import anthropic  # noqa: F401
    except ImportError:
        return {}, "自動読取ライブラリ(anthropic)が未導入のため、手入力で作成してください。"
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return {}, "自動読取(APIキー)が未設定のため、手入力で作成してください。"

    text_piece = ([{"type": "text", "text": req.text}]
                  if req.text and req.text.strip() else [])

    # 画像はそのまま1回だけ試す（分割不可）。
    if req.file_base64 and req.mime.startswith("image/"):
        piece = [{"type": "image", "source": {
            "type": "base64", "media_type": req.mime, "data": req.file_base64}}]
        try:
            data = _call_claude_json(req.doc_type, piece + text_piece)
        except Exception as e:  # noqa: BLE001
            return {}, f"画像の自動読取に失敗しました（{type(e).__name__}）。手入力で続行できます。"
        return (data, "") if data else ({}, "画像から項目を読み取れませんでした。手入力で続行できます。")

    # PDF
    if req.file_base64 and req.mime == "application/pdf":
        raw, size, pages, text = _pdf_stats(req.file_base64)

        def _doc(b64: str) -> list[dict[str, Any]]:
            return [{"type": "document", "source": {
                "type": "base64", "media_type": "application/pdf", "data": b64}}]

        small = size <= _MAX_DIRECT_BYTES and (pages is None or pages <= _MAX_DIRECT_PAGES)
        if small:
            try:
                data = _call_claude_json(req.doc_type, _doc(req.file_base64) + text_piece)
                if data:
                    return data, ""
            except Exception:  # noqa: BLE001 大きめ等で失敗 → 下のフォールバックへ
                pass

        # フォールバック1: 本文テキスト層が十分ならテキストだけ送る（軽い・限度回避）。
        good_text = text and len(text.strip()) >= max(500, 40 * (pages or 1))
        if good_text:
            try:
                data = _call_claude_json(
                    req.doc_type, [{"type": "text", "text": text[:120_000]}])
                if data:
                    note = "" if small else "資料が大きいため本文テキストから読み取りました（要確認）。"
                    return data, note
            except Exception:  # noqa: BLE001
                pass

        # フォールバック2: 頁分割して1バッチずつ処理し、取れた分を統合。
        if pages and pages > 0:
            per = max(1, min(_MAX_DIRECT_PAGES,
                             int(pages * _BATCH_MAX_BYTES / size) if size else pages))
            merged: dict[str, Any] = {}
            ok = fail = 0
            for start in range(0, pages, per):
                sub = _slice_pdf_b64(raw, start, start + per)
                if not sub:
                    fail += 1
                    continue
                try:
                    part = _call_claude_json(req.doc_type, _doc(sub))
                except Exception:  # noqa: BLE001
                    part = None
                if part:
                    merged = _merge_extracted(merged, part)
                    ok += 1
                else:
                    fail += 1
            if merged:
                note = "資料が大きいため分割して読み取りました（要確認）。" if fail == 0 else \
                    f"資料が大きく一部（{fail}ブロック）読み取れませんでした。取れた範囲を反映（要確認）。"
                return merged, note

        # テキスト層があるなら最後の望みで送る（分割も失敗した場合）。
        if text and text.strip():
            try:
                data = _call_claude_json(
                    req.doc_type, [{"type": "text", "text": text[:120_000]}])
                if data:
                    return data, "資料が重いためテキスト抽出で読み取りました（要確認）。"
            except Exception:  # noqa: BLE001
                pass
        return {}, "資料が重い/読みにくいため自動読取できませんでした。手入力で続行できます。"

    # テキストのみ
    if text_piece:
        try:
            data = _call_claude_json(req.doc_type, text_piece)
            return (data, "") if data else ({}, "テキストから項目を読み取れませんでした。手入力で続行できます。")
        except Exception as e:  # noqa: BLE001
            return {}, f"自動読取に失敗しました（{type(e).__name__}）。手入力で続行できます。"

    return {}, "読み取る資料（PDF/画像/テキスト）が指定されていません。"


@app.post("/extract", response_model=ExtractResp)
def extract(req: ExtractReq) -> ExtractResp:
    """AB書類を自動読取する。**どんな資料・重さでも 500 で止めない**：
    読み取れなければ空データ＋警告を返し、UI 側で手入力に切り替えられる。"""
    try:
        data, warning = _extract_robust(req)
    except BaseException as e:  # noqa: BLE001 想定外(panic含む)も握りつぶし手入力へ誘導（絶対に落とさない）
        data, warning = {}, f"自動読取で予期しない問題（{type(e).__name__}）。手入力で続行できます。"
    try:
        data = _normalize_extracted(data)
    except BaseException:  # noqa: BLE001 正規化失敗も無視（生データで返す）
        pass
    return ExtractResp(extracted=data, warning=warning)


def _normalize_extracted(data: dict[str, Any]) -> dict[str, Any]:
    """抽出結果の表記ゆれを正規化する（法令名→正式名称・用途地域→正式名称）。"""
    import horei_master

    horei = data.get("horei")
    if isinstance(horei, dict):
        laws = horei.get("other_horei")
        if isinstance(laws, list):
            horei["other_horei"] = [horei_master.normalize_horei(x) for x in laws]
        if horei.get("yoto"):
            horei["yoto"] = normalize_yoto(horei["yoto"])
    return data


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("BC_PORT", "8800")))
