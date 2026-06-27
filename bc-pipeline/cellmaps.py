"""本番ワークブックのセルマップ（テンプレ変種ごと）.

座標は同一テンプレの実例を差分比較して特定したもの（個人情報ではない）。
「間違いないように」のため、**確証の取れたセルのみ**を定義する。未確認のセルは
入れない（誤差し込みを避ける）。新しいテンプレ/シートは実例2通の差分（wb_diff.py）で
増やせる。

変種:
  36-1 … 土地建物（戸建）売主宅建業者用
  37-1 … 区分所有建物（敷地権）
  38-1 … 区分所有建物（非敷地権）
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import coordinate_to_tuple, get_column_letter

from bc_schema import YOTO_OPTIONS, normalize_yoto
from cellmap_grids import CHIIKI_CHIKU_MARKS, OTHER_HOREI_MARKS
from horei_master import normalize_horei
from keiyaku_schema import Keiyakusho
from juyojiko_schema import Juyojiko

CONTRACT_SHEET = "不動産売買契約書"
JUYOJIKO_SHEET = "重要事項説明書"


def _split_wareki(date_str: str | None) -> tuple[int, int, int] | None:
    """日付文字列を (令和年, 月, 日) に分解する。読めなければ None。

    対応: "令和7年1月1日" / "2025年4月10日" / "2025-04-10" / "2025/4/10"。
    令和 = 西暦 − 2018（令和元年=2019）。
    """
    if not date_str:
        return None
    s = str(date_str).strip()
    m = re.search(r"令和\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", s)
    if m:
        return int(m[1]), int(m[2]), int(m[3])
    m = re.search(r"(\d{4})\s*年\s*(\d+)\s*月\s*(\d+)\s*日", s)
    if m:
        return int(m[1]) - 2018, int(m[2]), int(m[3])
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return int(m[1]) - 2018, int(m[2]), int(m[3])
    return None


def _date_cells(date_str: str | None, y: str, mo: str, d: str) -> dict[str, int]:
    """日付を 令和年/月/日 の3セルに分けた {coord: 数値} を返す（読めなければ空）。"""
    parsed = _split_wareki(date_str)
    if not parsed:
        return {}
    return {y: parsed[0], mo: parsed[1], d: parsed[2]}


def _split_era_date(s: str | None) -> tuple[str, int, int, int] | None:
    """日付を (元号, 年, 月, 日) に分解。令和/平成/昭和・西暦に対応。読めなければ None。"""
    if not s:
        return None
    s = str(s).strip()
    m = re.search(r"(令和|平成|昭和)\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", s)
    if m:
        return m[1], int(m[2]), int(m[3]), int(m[4])
    m = re.search(r"(\d{4})\s*[年/-]\s*(\d{1,2})\s*[月/-]\s*(\d{1,2})", s)
    if m:
        y = int(m[1])
        era, yy = ("令和", y - 2018) if y >= 2019 else ("平成", y - 1988)
        return era, yy, int(m[2]), int(m[3])
    return None


def _kakunin_cells(date: str | None, num: str | None, b: str,
                   era: str, y: str, mo: str, d: str, num_c: str) -> dict[str, Any]:
    """建築確認/検査済証の (有無・元号年月日・番号) セルを返す。"""
    out: dict[str, Any] = {}
    if num:
        out[num_c] = num
        out[b] = ON
    p = _split_era_date(date)
    if p:
        out[era], out[y], out[mo], out[d] = p
    return out


def _split_chiban(shozai: str | None) -> tuple[str, str | None, str | None]:
    """所在地を (所在の前置き, 番, 番地) に分解する。

    例 "○○市○○町12番5" → ("○○市○○町", "12", "5")。
    末尾に「N番M」「N番地M」「N番」が無ければ (全体, None, None)。
    """
    if not shozai:
        return "", None, None
    s = str(shozai).strip()
    m = re.search(r"(\d+)\s*番(?:地)?\s*(\d+)?\s*$", s)
    if not m:
        return s, None, None
    return s[:m.start()].strip(), m.group(1), m.group(2)


def _chiban_cells(shozai: str | None, prefix: str, ban: str, banchi: str) -> dict[str, Any]:
    """土地の所在を 所在/番/番地 の3セルに分けた {coord: 値} を返す。"""
    if not shozai:
        return {}
    pre, b, bc = _split_chiban(shozai)
    out: dict[str, Any] = {prefix: pre}
    if b:
        out[ban] = b
    if bc:
        out[banchi] = bc
    return out

ON, OFF = "■", "□"

# 区域区分のチェックセル（変種別）
KUIKI_MARKS = {
    "36-1": {"市街化区域": "T331", "市街化調整区域": "AA331",
             "区域区分のされていない区域": "AJ331"},
    "区分": {"市街化区域": "T335", "市街化調整区域": "AA335",
             "区域区分のされていない区域": "AJ335"},
}
# 都市計画区域内/外のチェックセル（区域区分の上位。変種別）
TOSHIKEIKAKU_MARKS = {
    "36-1": {"都市計画区域内": "J331", "都市計画区域外": "J333"},
    "区分": {"都市計画区域内": "J335", "都市計画区域外": "J337"},
}
# 用途地域14選択肢のチェックセル（YOTO_OPTIONS の順）。変種別。
YOTO_MARKS = {
    "36-1": ["C356", "C358", "C360", "C362", "C364", "R356", "R358",
             "R360", "R362", "R364", "AG356", "AG358", "AG360", "AG362"],
    "区分": ["C360", "C362", "C364", "C366", "C368", "R360", "R362",
             "R364", "R366", "R368", "AG360", "AG362", "AG364", "AG366"],
}
# 防火関係（相互排他）。変種別。
BOKA_MARKS = {
    "36-1": {"防火地域": "C368", "準防火地域": "C370", "新たな防火規制区域": "C372"},
    "区分": {"防火地域": "C372", "準防火地域": "C374", "新たな防火規制区域": "C376"},
}
# 設備の種別チェックセル（変種別。実例で特定）
SUIDOU_MARKS = {
    "36-1": {"公営水道": "G643", "私営水道": "G645", "井戸": "G647"},
    "区分": {"公営水道": "G647", "私営水道": "G649", "井戸": "G651"},
}
GAS_MARKS = {
    "36-1": {"都市ガス": "G656", "個別プロパン": "G660", "集中プロパン": "G662"},
    "区分": {"都市ガス": "G660", "個別プロパン": "G664", "集中プロパン": "G666"},
}
OSUI_MARKS = {
    "36-1": {"公共下水": "G664", "個別浄化槽": "G666", "集中浄化槽": "G668", "汲取式": "G670"},
    "区分": {"公共下水": "G668", "個別浄化槽": "G670", "集中浄化槽": "G672", "汲取式": "G674"},
}
ZASSUI_MARKS = {
    "36-1": {"公共下水": "G674", "個別浄化槽": "G676", "集中浄化槽": "G678",
             "側溝等": "G680", "浸透式": "G682"},
    "区分": {"公共下水": "G678", "個別浄化槽": "G680", "集中浄化槽": "G682",
             "側溝等": "G684", "浸透式": "G686"},
}
DENRYOKU_CELL = {"36-1": "G652", "区分": "G653"}


def _setsubi_values(variant: str, sd: Any, biko_coord: str | None,
                    biko_fallback: str | None) -> dict[str, Any]:
    """設備の種別チェック＋電力会社＋備考の差込値を返す。"""
    out: dict[str, Any] = {}
    out.update(_checkbox(SUIDOU_MARKS[variant], _g(sd, "suidou")))
    out.update(_checkbox(GAS_MARKS[variant], _g(sd, "gas")))
    out.update(_checkbox(OSUI_MARKS[variant], _g(sd, "osui")))
    out.update(_checkbox(ZASSUI_MARKS[variant], _g(sd, "zassui")))
    out[DENRYOKU_CELL[variant]] = _g(sd, "denryoku")
    if biko_coord:
        out[biko_coord] = _g(sd, "biko") or biko_fallback
    return out

# 建築基準法第22条区域（独立チェック）。変種別。
NIJUNI_MARK = {"36-1": "C374", "区分": "C378"}
# 高度地区（独立チェック）。変種別。
KODO_MARK = {"36-1": "C376", "区分": "C380"}


def _norm_boka(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    if "新た" in s:
        return "新たな防火規制区域"
    if "準防火" in s:
        return "準防火地域"
    if "防火" in s:
        return "防火地域"
    return None


def _norm_kuiki(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    if "調整" in s:
        return "市街化調整区域"
    if "されていない" in s or "非線引" in s or "未線引" in s:
        return "区域区分のされていない区域"
    if "市街化区域" in s:
        return "市街化区域"
    return s


def _shakuchi_lines(bc: Any) -> list[str]:
    """借地条件を1行ずつの可読テキストにする（借地物件のときのみ）。

    専用の借地説明書シートは実データ未照合のため差し込まないが、
    情報を失わないよう標準重説のⅤ備考へ転記する。
    """
    sh = getattr(bc, "shakuchi", None)
    if sh is None:
        return []
    parts: list[str] = []
    shurui = _g(sh, "shakuchiken_shurui")
    if shurui:
        parts.append(f"借地権の種類: {shurui}")
    sonzoku = _g(sh, "sonzoku_kikan")
    if sonzoku:
        parts.append(f"存続期間: {sonzoku}")
    jidai = _g(sh, "jidai_kingaku")
    if jidai is not None:
        tani = _g(sh, "jidai_tani") or ""
        parts.append(f"地代: {tani}{jidai:,}円")
    for label, key in (("地代支払方法", "jidai_shiharai"), ("地代改定", "jidai_kaitei"),
                       ("更新料", "koshin_ryo"), ("譲渡承諾", "joto_shodaku"),
                       ("建築制限", "kenchiku_seigen")):
        v = _g(sh, key)
        if v:
            parts.append(f"{label}: {v}")
    jusho = _g(sh, "teichi_shoyusha_jusho")
    shimei = _g(sh, "teichi_shoyusha_shimei")
    if jusho or shimei:
        parts.append(f"底地所有者（地主）: {(jusho or '')} {(shimei or '')}".strip())
    biko = _g(sh, "biko")
    if biko:
        parts.append(f"借地備考: {biko}")
    return ["【借地条件】"] + parts if parts else []


def _biko_text(bc: Any) -> str | None:
    """Ⅴ備考の自由記述（容認事項＋特約＋借地条件）を1セル分のテキストにまとめる。"""
    lines = list(getattr(bc, "yonin_jiko", None) or []) + \
        list(getattr(bc, "tokuyaku", None) or []) + \
        _shakuchi_lines(bc)
    return "\n".join(lines) if lines else None


def _toggle(false_coord: str, true_coord: str, val: bool | None) -> dict[str, str]:
    """2択トグル（外/内・有/無 等）。val=True→true側■、False→false側■。None→空。"""
    if val is None:
        return {}
    return {true_coord: ON, false_coord: OFF} if val else {false_coord: ON, true_coord: OFF}


def _checkbox(option_to_coord: dict[str, str], selected: str | None) -> dict[str, str]:
    """選択肢→セルの対応から {coord: ■/□} を返す。selected が無ければ空（テンプレ非改変）。"""
    if not selected:
        return {}
    return {coord: (ON if opt == selected else OFF)
            for opt, coord in option_to_coord.items()}


# その他の地域地区(格子)の先頭5ゾーンは防火/22条/高度の専用フィールド
# (boka/nijuni_jo/kodo_chiku)が同一セル(C368〜C376系)を扱う。格子側から除外し、
# 格子の□初期化が専用チェックの■を打ち消す二重管理バグを防ぐ。
_CHIIKI_DEDICATED = ("防火地域", "準防火地域", "新たな防火規制区域",
                     "建築基準法第22条区域", "高度地区")


def _chiiki_chiku_marks(variant_key: str) -> dict[str, str]:
    """その他の地域地区の格子マップから、専用フィールドが扱う先頭5ゾーンを除いたものを返す。"""
    return {z: c for z, c in CHIIKI_CHIKU_MARKS[variant_key].items()
            if z not in _CHIIKI_DEDICATED}


def _horei_grid(marks: dict[str, str], selected: list[str] | None) -> dict[str, str]:
    """法令/地域地区のチェック格子に ■/□ を差し込む。

    `selected` が空なら何もしない（_checkbox と同じく非改変）。データがあるときは格子全体を
    □ で初期化（異物件WB流用時の残留防止）してから、該当する枠だけ ■ にする。
    様式に無い項目（例: 37-1 様式に無い生物多様性増進法）は黙ってスキップする。
    """
    if not selected:
        return {}
    out: dict[str, str] = {coord: OFF for coord in marks.values()}
    for name in selected:
        coord = marks.get(normalize_horei(name))
        if coord:
            out[coord] = ON
    return out


# 日影規制 有/無 チェック（有, 無）。変種別。構造化欄の種別・時間は自由文字列から
# 確実に取れないため差し込まない（捏造回避）。
NISSHIDO_MARKS = {
    "36-1": ("L412", "O412"),
    "37-1": ("L416", "O416"),
    "38-1": ("L416", "O416"),
}


def _nisshido_cells(variant: str, nisshido: str | None) -> dict[str, str]:
    """日影規制の 有/無 チェックを返す。値が無ければ非改変。"""
    if not nisshido:
        return {}
    yes, no = NISSHIDO_MARKS[variant]
    s = str(nisshido)
    is_none = any(k in s for k in ("無", "なし", "対象外", "指定なし", "非該当"))
    return {no: ON, yes: OFF} if is_none else {yes: ON, no: OFF}


def _juyojiko_checkboxes(variant_key: str, h: Any) -> dict[str, str]:
    """区域区分・用途地域・地域地区のチェック差込値をまとめて返す。"""
    out: dict[str, str] = {}
    # 都市計画区域内/外（区域区分の上位チェック）
    tk = _g(h, "toshikeikaku_kuiki")
    if tk:
        sel = "都市計画区域外" if "外" in str(tk) else "都市計画区域内"
        out.update(_checkbox(TOSHIKEIKAKU_MARKS[variant_key], sel))
    out.update(_checkbox(KUIKI_MARKS[variant_key], _norm_kuiki(_g(h, "kuiki_kubun"))))
    yoto_map = dict(zip(YOTO_OPTIONS, YOTO_MARKS[variant_key]))
    out.update(_checkbox(yoto_map, normalize_yoto(_g(h, "yoto"))))
    # 防火関係（相互排他）
    out.update(_checkbox(BOKA_MARKS[variant_key], _norm_boka(_g(h, "boka"))))
    # 建築基準法第22条区域（独立。True/False が分かるときだけ）
    nijuni = _g(h, "nijuni_jo")
    if nijuni is not None:
        out[NIJUNI_MARK[variant_key]] = ON if nijuni else OFF
    # 高度地区（独立。値があるときだけ ■）
    if _g(h, "kodo_chiku"):
        out[KODO_MARK[variant_key]] = ON
    return out


# 違約金（「2.売買代金の N%」を選択＋%値）。(選択肢2□, %値, 選択肢1□, 選択肢3□)。変種別。
IYAKUKIN_CELLS = {
    "36-1": ("O1008", "W1008", "G1008", "AD1008"),
    "区分": ("O1256", "W1256", "G1256", "AD1256"),
}
# 担保責任の措置（1.講じる / 2.講じない）。変種別。
TANPO_CELLS = {"36-1": ("T1078", "Z1078"), "区分": ("T1328", "Z1328")}
# 契約書 表紙「違約金の額」。重説と同じ (選択肢2□, %値, 選択肢1□, 選択肢3□)。
# 37-1・38-1 は契約書レイアウト同一のため「区分」で共用（実例で確認）。
KEIYAKU_IYAKUKIN_CELLS = {
    "36-1": ("X65", "AF65", "P65", "AM65"),
    "区分": ("X70", "AF70", "P70", "AM70"),
}


def _iyakukin_select(cells: tuple[str, str, str, str], iw: int | None) -> dict[str, Any]:
    """違約金「2.売買代金の N%相当額」を選択し%を差し込む。cells=(選択肢2□,%値,選択肢1□,選択肢3□)。

    iw が None のときは非改変（テンプレ既定の選択を温存）。
    """
    if iw is None:
        return {}
    opt2, pct, opt1, opt3 = cells
    return {opt1: OFF, opt2: ON, opt3: OFF, pct: iw}


def _joken_cells(variant_key: str, jk: Any) -> dict[str, Any]:
    """Ⅱ取引条件のうち違約金%・担保責任の措置をチェック/値で差し込む。"""
    out: dict[str, Any] = {}
    out.update(_iyakukin_select(IYAKUKIN_CELLS[variant_key], _g(jk, "iyakukin_wariai")))
    tp = _g(jk, "tanpo_sekinin")
    if tp:
        kouji, kouji_nai = TANPO_CELLS[variant_key]  # 講じる, 講じない
        s = str(tp)
        if "講じない" in s:
            out[kouji] = OFF
            out[kouji_nai] = ON
        elif "講じる" in s:
            out[kouji] = ON
            out[kouji_nai] = OFF
    return out


def _g(obj: Any, *path: str) -> Any:
    for p in path:
        if obj is None:
            return None
        obj = getattr(obj, p, None) if not isinstance(obj, dict) else obj.get(p)
    return obj


# 契約書 表紙の追加記入欄（変種別）。区分=37-1/38-1 は契約書レイアウト共通。
# 値=単一セル、tuple=分割セル（日付は令和年/月/日、締結日は元号/年/月/日、有無は有/無）。
KEIYAKU_OMOTE_CELLS = {
    "36-1": {
        "uchikin1": "AE55", "uchikin1_date": ("S55", "W55", "AA55"),
        "hikiwatashi": "AE61", "seisan": ("S63", "W63", "AA63"),
        "loan_umu": ("Q67", "U67"), "loan_kingaku": "AE71",
        "loan_kaijo": ("AH81", "AL81", "AP81"),
        "gyosha_shozai": "P135", "gyosha_shomei": "P137", "gyosha_daihyo": "P139",
        "torikiishi": "P143", "keiyaku_date": ("AI130", "AL130", "AP130", "AT130"),
    },
    "区分": {
        "uchikin1": "AE60", "uchikin1_date": ("S60", "W60", "AA60"),
        "hikiwatashi": "AE66", "seisan": ("S68", "W68", "AA68"),
        "loan_umu": ("Q72", "U72"), "loan_kingaku": "AE76",
        "loan_kaijo": ("AH86", "AL86", "AP86"),
        "gyosha_shozai": "P140", "gyosha_shomei": "P142", "gyosha_daihyo": "P144",
        "torikiishi": "P148", "keiyaku_date": ("AI135", "AL135", "AP135", "AT135"),
    },
}


def _keiyaku_omote_values(variant_key: str, bc: Keiyakusho) -> dict[str, Any]:
    """契約書 表紙の追加欄（内金①・引渡日・公租公課起算日・融資・業者/取引士・締結日）を差し込む。"""
    m = KEIYAKU_OMOTE_CELLS[variant_key]
    d, g, t = bc.daikin, bc.gyosha, bc.torikiishi
    out: dict[str, Any] = {
        m["uchikin1"]: _g(d, "uchikin1"),
        m["hikiwatashi"]: _g(bc, "hikiwatashi_date"),
        m["loan_kingaku"]: _g(bc, "loan_kingaku"),
        m["gyosha_shozai"]: _g(g, "shozai"),
        m["gyosha_shomei"]: _g(g, "shomei"),
        m["gyosha_daihyo"]: _g(g, "daihyo"),
        m["torikiishi"]: _g(t, "shimei"),
    }
    out.update(_date_cells(_g(d, "uchikin1_date"), *m["uchikin1_date"]))
    out.update(_date_cells(_g(bc, "seisan_kisanbi"), *m["seisan"]))
    out.update(_date_cells(_g(bc, "loan_kaijo_date"), *m["loan_kaijo"]))
    # 融資利用の有無（有/無トグル。True→有■）
    lt = _g(bc, "loan_tokuyaku")
    if lt is not None:
        umu_yes, umu_no = m["loan_umu"]
        out.update(_toggle(umu_no, umu_yes, lt))
    # 契約締結日（元号/年/月/日）
    kd = _split_era_date(_g(bc, "keiyaku_date"))
    if kd:
        era_c, y_c, mo_c, d_c = m["keiyaku_date"]
        out[era_c], out[y_c], out[mo_c], out[d_c] = kd
    return out


def _keiyaku_omote_clear(variant_key: str) -> list[str]:
    """表紙追加欄の全セル座標（分割セル含む）を返す。未充当時も残留しないようクリア対象にする。"""
    out: list[str] = []
    for v in KEIYAKU_OMOTE_CELLS[variant_key].values():
        out.extend(v if isinstance(v, tuple) else [v])
    return out


def _build_keiyaku_36_1(bc: Keiyakusho) -> tuple[dict[str, Any], list[str]]:
    """36-1（土地建物）契約書シートの (差込値, 追加クリアセル) を返す。

    差分検証済みセル:
      当事者 E123(売主)/AB123(買主)、代金 AE45/内訳AE47・AE49・AE51/手付AE53/残代金AE59、
      土地 F11(所在)/AF11(地目)/AL11(地積)、建物 I23(所在)/AK23(家屋番号)/AE25(構造)/AR29(床面積)、
      特記 D31。
    """
    d = bc.daikin
    f = bc.fudosan
    tochi = _g(f, "tochi")
    tate = _g(f, "tatemono")
    values: dict[str, Any] = {
        "E123": _g(bc, "urinushi", "name"),
        "AB123": _g(bc, "kainushi", "name"),
        "AE45": _g(d, "baibai_daikin"),
        "AE47": _g(d, "tochi_kakaku"),
        "AE49": _g(d, "tatemono_kakaku"),
        "AE51": _g(d, "shohizei"),
        "AE53": _g(d, "tetsuke"),
        "AE59": _g(d, "zankin"),
        "F11": _g(tochi, "shozai"),
        "AF11": _g(tochi, "chimoku"),
        "AL11": _g(tochi, "chiseki_toki"),
        "I23": _g(tate, "shozai") or _g(tochi, "shozai"),
        "AK23": _g(tate, "kaoku_bango"),
        "AE25": _g(tate, "kozo"),
        "AR29": _g(tate, "yukamenseki"),
        "D31": "\n".join(bc.tokuyaku) if bc.tokuyaku else None,
    }
    # 土地所在の分割差込（所在 F11 / 番 X11 / 番地 AC11）
    values.update(_chiban_cells(_g(tochi, "shozai"), "F11", "X11", "AC11"))
    # 日付の分割差込（令和年/月/日）。残代金支払日 S59・融資承認取得期日 O71（実例検証済み）
    values.update(_date_cells(_g(d, "zankin_date"), "S59", "W59", "AA59"))
    values.update(_date_cells(_g(bc, "loan_shonin_date"), "O71", "S71", "W71"))
    # 表紙「違約金の額」= 売買代金の N%（重説 Ⅱ取引条件と整合）
    values.update(_iyakukin_select(KEIYAKU_IYAKUKIN_CELLS["36-1"], _g(d, "iyakukin_wariai")))
    # 表紙の追加欄（内金①・引渡日・公租公課起算日・融資・業者/取引士・締結日）
    values.update(_keiyaku_omote_values("36-1", bc))
    # 旧案件の値が残らないようクリアする（差込しない地番・日付・備考の分割セル）
    clear_extra = ["X11", "AC11", "S59", "W59", "AA59",
                   "O71", "S71", "W71", "B100"] + _keiyaku_omote_clear("36-1")
    return values, clear_extra


def _build_keiyaku_kubun(bc: Keiyakusho) -> tuple[dict[str, Any], list[str]]:
    """区分（37-1=敷地権 / 38-1=非敷地権）契約書シートの (差込値, 追加クリアセル)。

    37-1・38-1 は契約書シートのレイアウトが同一（実例で確認）。
    実例検証済みセル:
      当事者 E128(売主)/AB128(買主)、
      一棟 I11(所在)/I13(名称)/I15(構造)/AO15(延床)、
      専有 G19(家屋番号)/AA19(建物名称)/AQ19(種類)/G21(構造)/AP21(床面積)、
      敷地権 D29(所在)/R29(地番)/Z29(地目)/AF29(地積)/AL29(種類)/AR29(割合)、
      代金 AE54(売買代金)/AE58(手付)/AE64(残代金)/AE66(引渡日)。
    """
    d = bc.daikin
    f = bc.fudosan
    se = _g(f, "senyuu")
    sk = (f.shikichiken[0] if (f and f.shikichiken) else None)
    values: dict[str, Any] = {
        "E128": _g(bc, "urinushi", "name"),
        "AB128": _g(bc, "kainushi", "name"),
        "AE54": _g(d, "baibai_daikin"),
        "AE58": _g(d, "tetsuke"),
        "AE64": _g(d, "zankin"),
        "AE66": _g(bc, "hikiwatashi_date"),
        "I11": _g(f, "ittou_shozai"),
        "I13": _g(f, "ittou_meisho"),
        "I15": _g(f, "ittou_kozo"),
        "AO15": _g(f, "ittou_enshoumenseki"),
        "G19": _g(se, "kaoku_bango"),
        "AA19": _g(se, "meisho"),
        "AQ19": _g(se, "shurui"),
        "G21": _g(se, "kozo"),
        "AP21": _g(se, "yukamenseki"),
        "D29": _g(sk, "shozai"),
        "R29": _g(sk, "chiban"),
        "Z29": _g(sk, "chimoku"),
        "AF29": _g(sk, "chiseki"),
        "AL29": _g(sk, "shikichiken_shurui"),
        "AR29": _g(sk, "wariai"),
    }
    # 表紙「違約金の額」= 売買代金の N%（重説 Ⅱ取引条件と整合。37-1/38-1 共通レイアウト）
    values.update(_iyakukin_select(KEIYAKU_IYAKUKIN_CELLS["区分"], _g(d, "iyakukin_wariai")))
    # 表紙の追加欄（内金①・引渡日・公租公課起算日・融資・業者/取引士・締結日）
    values.update(_keiyaku_omote_values("区分", bc))
    # 旧案件の金額（消費税・内金②）をクリア
    clear_extra = ["AE56", "AE62"] + _keiyaku_omote_clear("区分")
    return values, clear_extra


# 変種 → 契約書ビルダー
KEIYAKU_BUILDERS = {
    "36-1": _build_keiyaku_36_1,
    "37-1": _build_keiyaku_kubun,
    "38-1": _build_keiyaku_kubun,
}


def build_keiyaku(variant: str, bc: Keiyakusho) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    """(sheet_values, sheet_clear) を返す。wb_fill.fill_workbook にそのまま渡せる。"""
    builder = KEIYAKU_BUILDERS.get(variant)
    if builder is None:
        raise KeyError(f"未対応のテンプレ変種: {variant}（対応: {list(KEIYAKU_BUILDERS)}）")
    values, clear_extra = builder(bc)
    clear = list(values.keys()) + clear_extra
    return {CONTRACT_SHEET: values}, {CONTRACT_SHEET: clear}


def _build_juyojiko_36_1(bc: Juyojiko, variant: str = "36-1") -> tuple[dict[str, Any], list[str]]:
    """36-1（土地建物）重説シートの (差込値, 追加クリアセル) を返す。

    差分・実例検証済みセル:
      売主 F261(住所)/F263(氏名)/F265(備考)、買主 F7、
      土地 D194(所在)/AF194(地目)/AL194(地積)、
      建物 G236(所在)/AK236(家屋番号)/G238(住居表示)/AC240(構造)/AR244(床面積)、
      指定建蔽率 Q384、指定容積率 Q398。
    """
    f = bc.fudosan
    tochi = _g(f, "tochi")
    tate = _g(f, "tatemono")
    h = bc.horei
    values: dict[str, Any] = {
        "F7": _g(bc, "kainushi", "name"),
        "F261": _g(bc, "urinushi", "address"),
        "F263": _g(bc, "urinushi", "name"),
        "F265": "\n".join(bc.tokuyaku) if bc.tokuyaku else None,
        "AF194": _g(tochi, "chimoku"),
        "AL194": _g(tochi, "chiseki_toki"),
        "G236": _g(tate, "shozai") or _g(tochi, "shozai"),
        "AK236": _g(tate, "kaoku_bango"),
        "G238": _g(f, "jukyo_hyoji"),
        "AC240": _g(tate, "kozo"),
        "AR244": _g(tate, "yukamenseki"),
        "Q384": _g(h, "kenpei"),
        "Q398": _g(h, "yoseki"),
    }
    # 土地所在の分割差込（所在 D194 / 番 X194 / 番地 AC194）
    values.update(_chiban_cells(_g(tochi, "shozai"), "D194", "X194", "AC194"))
    # 宅建業者・取引士欄（BC側媒介。案件マスタ由来。検証済みセル）
    g = bc.gyosha
    t = bc.torikiishi
    values["H23"] = _g(g, "shozai")          # 業者 主たる事務所の所在地
    values["AF31"] = _g(g, "daihyo")         # 業者 代表者氏名
    values["H35"] = _g(t, "shimei")          # 取引士 氏名
    values["H39"] = _g(t, "jimusho_shozai")  # 取引士 事務所所在地
    values["R406"] = _g(h, "shikichi_saitei")  # 敷地面積の最低限度
    values["AK238"] = _g(f, "fuzoku_tatemono")  # 附属建物の有無
    # 接面道路（方向 D440 / 幅員 X440 / 接道長さ AE440 / 備考 AL438）
    values["D440"] = _g(h, "doro_hoko")
    values["X440"] = _g(h, "doro_haba")
    values["AE440"] = _g(h, "doro_setsudo")
    values["AL438"] = _g(h, "doro")
    # 設備（飲用水・ガス・排水の種別チェック＋電力会社・備考）
    values.update(_setsubi_values("36-1", bc.setsubi_detail, "B695", bc.setsubi))
    # 災害区域・調査（外/内・有/無トグル。AB引継ぎ）
    sg = bc.saigai
    values.update(_toggle("Z795", "AD795", _g(sg, "zosei_bosai")))      # 造成宅地防災
    values.update(_toggle("Z800", "AD800", _g(sg, "dosha_keikai")))     # 土砂災害警戒
    values.update(_toggle("Z802", "AD802", _g(sg, "dosha_tokubetsu")))  # 土砂特別警戒
    values.update(_toggle("Z807", "AD807", _g(sg, "tsunami_keikai")))   # 津波警戒
    values.update(_toggle("Z809", "AD809", _g(sg, "tsunami_tokubetsu")))  # 津波特別警戒
    values.update(_toggle("V844", "R844", _g(sg, "taishin_shindan")))   # 耐震診断 有/無
    if _g(sg, "sekimen_kiroku") is not None:
        values["F831"] = ON if _g(sg, "sekimen_kiroku") else OFF        # 石綿記録の有無
    # 水害ハザード（洪水 W814/AA814・内水 AM814/AQ814・高潮 W816/AA816）
    values.update(_toggle("AA814", "W814", _g(sg, "kozui")))
    values.update(_toggle("AQ814", "AM814", _g(sg, "naisui")))
    values.update(_toggle("AA816", "W816", _g(sg, "takashio")))
    # 建築確認・検査済証（有無・元号年月日・番号）
    kk = bc.kakunin
    values.update(_kakunin_cells(_g(kk, "kenchiku_date"), _g(kk, "kenchiku_bango"),
                                 "B780", "R780", "U780", "Y780", "AC780", "AG780"))
    values.update(_kakunin_cells(_g(kk, "kensa_date"), _g(kk, "kensa_bango"),
                                 "B782", "R782", "U782", "Y782", "AC782", "AG782"))
    # 登記記録の権利（所有者・乙区。三為=元所有者のまま＝AB引継ぎ）
    tk = bc.touki
    values["L288"] = _g(tk, "tochi_shoyusha_jusho")
    values["L290"] = _g(tk, "tochi_shoyusha_shimei")
    values["L296"] = _g(tk, "tochi_otsuku")
    values["L308"] = _g(tk, "tatemono_shoyusha_jusho")
    values["L310"] = _g(tk, "tatemono_shoyusha_shimei")
    values["L316"] = _g(tk, "tatemono_otsuku")
    values["F277"] = bc.senyuusha_uchi            # 占有に関する事項
    values["B250"] = _g(f, "fuzoku_tatemono_detail")  # 附属建物の詳細
    values["O818"] = _g(h, "suigai_shozai")       # 水害ハザード 所在地の説明
    values["B891"] = _g(bc, "seisan_biko")        # 公租公課の清算 備考
    values["B1196"] = _biko_text(bc)              # Ⅴ備考（容認事項＋特約）
    values.update(_juyojiko_checkboxes("36-1", h))  # 区域区分・用途地域・都市計画区域の■/□
    values.update(_joken_cells("36-1", bc.joken))   # 違約金%・担保責任の措置
    # 地積（実測。戸建のみ本欄あり）・建築時期（元号/年/月/日に分割）
    values["G206"] = _g(tochi, "chiseki_jissoku")
    ck = _split_era_date(_g(tate, "chikujiki"))
    if ck:
        values["H246"], values["K246"], values["O246"], values["S246"] = ck
    # 日影規制（有/無）・その他の地域地区(22)・都計法外の法令(61)のチェック格子
    values.update(_nisshido_cells("36-1", _g(h, "nisshido")))
    values.update(_horei_grid(_chiiki_chiku_marks("36-1"), _g(h, "chiiki_chiku")))
    values.update(_horei_grid(OTHER_HOREI_MARKS["36-1"], _g(h, "other_horei")))
    # 旧案件の値が残らないようクリア（差込しない地番・床面積の分割セル）
    clear_extra = ["X194", "AC194", "P242", "X242"]
    return values, clear_extra


def _build_juyojiko_kubun(bc: Juyojiko, variant: str = "37-1") -> tuple[dict[str, Any], list[str]]:
    """区分（37-1 / 38-1）重説シートの (差込値, 追加クリアセル)。

    37-1・38-1 は重説シートのレイアウトが同一（実例で確認）。
    実例検証済みセル:
      売主 F265(住所)/F267(氏名)、買主 F7、
      一棟 I194(名称)/L201(所在)/L205(延床)、専有 L207(家屋番号)/AL207(建物名称)、
      指定建蔽率 Q388、指定容積率 Q402、
      修繕積立金 月額 L864 / 積立累計 U868 / 滞納 V872、管理費 月額 L884 / 滞納 V888、
      売買代金 H1116、手付金 V1129。
    """
    f = bc.fudosan
    se = _g(f, "senyuu")
    h = bc.horei
    k = bc.kanri
    j = bc.joken
    values: dict[str, Any] = {
        "F7": _g(bc, "kainushi", "name"),
        "F265": _g(bc, "urinushi", "address"),
        "F267": _g(bc, "urinushi", "name"),
        "I194": _g(f, "ittou_meisho"),
        "L201": _g(f, "ittou_shozai"),
        "L205": _g(f, "ittou_enshoumenseki"),
        "L207": _g(se, "kaoku_bango"),
        "AL207": _g(se, "meisho"),
        "Q388": _g(h, "kenpei"),
        "Q402": _g(h, "yoseki"),
        "L864": _g(k, "shuzen_getsugaku"),
        "U868": _g(k, "shuzen_tsumitate"),
        "V872": _g(k, "shuzen_taino"),
        "L884": _g(k, "kanrihi_getsugaku"),
        "V888": _g(k, "kanrihi_taino"),
        "K900": _g(k, "kanri_kumiai"),       # 管理組合の名称
        "M902": _g(k, "kanri_keitai"),       # 管理形態（全部委託 等）
        "K904": _g(k, "kanri_itakusaki"),    # 管理委託先
        "L747": _g(k, "yoto_seigen"),        # 専有部分の用途制限
        "L751": _g(k, "pet_seigen"),         # ペット飼育制限
        "H1116": _g(j, "baibai_daikin"),
        "V1129": _g(j, "tetsuke"),
    }
    # 宅建業者・取引士欄（BC側媒介。案件マスタ由来。区分は代表者が H31）
    g = bc.gyosha
    t = bc.torikiishi
    values["H23"] = _g(g, "shozai")
    values["H31"] = _g(g, "daihyo")
    values["H35"] = _g(t, "shimei")
    values["H39"] = _g(t, "jimusho_shozai")
    # 災害区域（外/内トグル。AB引継ぎ。区分の行位置）
    sg = bc.saigai
    values.update(_toggle("Z1048", "AD1048", _g(sg, "dosha_keikai")))
    values.update(_toggle("Z1050", "AD1050", _g(sg, "dosha_tokubetsu")))
    values.update(_toggle("Z1055", "AD1055", _g(sg, "tsunami_keikai")))
    values.update(_toggle("Z1057", "AD1057", _g(sg, "tsunami_tokubetsu")))
    # 設備（区分の種別チェック＋電力会社）
    values.update(_setsubi_values("区分", bc.setsubi_detail, None, None))
    # 建築確認・検査済証（区分の行位置）
    kk = bc.kakunin
    values.update(_kakunin_cells(_g(kk, "kenchiku_date"), _g(kk, "kenchiku_bango"),
                                 "B1028", "R1028", "U1028", "Y1028", "AC1028", "AG1028"))
    values.update(_kakunin_cells(_g(kk, "kensa_date"), _g(kk, "kensa_bango"),
                                 "B1030", "R1030", "U1030", "Y1030", "AC1030", "AG1030"))
    # 水害ハザード（区分の行位置）
    values.update(_toggle("AA1062", "W1062", _g(sg, "kozui")))
    values.update(_toggle("AQ1062", "AM1062", _g(sg, "naisui")))
    values.update(_toggle("AA1064", "W1064", _g(sg, "takashio")))
    # 登記の権利（区分: 専有=建物 L292/L294/L300、敷地権=土地 L312/L314/L320）＋占有
    tk = bc.touki
    values["L292"] = _g(tk, "tatemono_shoyusha_jusho")
    values["L294"] = _g(tk, "tatemono_shoyusha_shimei")
    values["L300"] = _g(tk, "tatemono_otsuku")
    values["L312"] = _g(tk, "tochi_shoyusha_jusho")
    values["L314"] = _g(tk, "tochi_shoyusha_shimei")
    values["L320"] = _g(tk, "tochi_otsuku")
    values["F281"] = bc.senyuusha_uchi
    # 占有者の住所(F277)・氏名(F279)は区分レイアウトでは個別セル。BC側は占有概要を
    # F281 にまとめて記載するため、テンプレ(他号室)の占有者PIIが残らないよう必ずクリアする。
    clears_extra = ["F277", "F279"]
    # 案件固有だが当方スキーマに項目が無いセル（委託先の登録番号・別欄の委託先表記・
    # Ⅳ添付書類の書類名リスト）は、他物件テンプレ流用時にデータが残らないよう必ずクリア。
    clears_extra += ["AQ908", "R1075"]
    clears_extra += [f"E{r}" for r in range(1425, 1444, 2)]   # 添付書類 左列
    clears_extra += [f"AC{r}" for r in range(1425, 1444, 2)]  # 添付書類 右列
    values["O1066"] = _g(h, "suigai_shozai")      # 水害ハザード 所在地の説明
    # 区分は 容認事項(B1366)と特約(B1449)が分かれている
    values["B1366"] = "\n".join(bc.yonin_jiko) if bc.yonin_jiko else None
    values["B1449"] = "\n".join(bc.tokuyaku) if bc.tokuyaku else None
    values.update(_juyojiko_checkboxes("区分", h))  # 区域区分・用途地域・都市計画区域の■/□
    values.update(_joken_cells("区分", j))          # 違約金%・担保責任の措置
    # 日影規制（有/無）・その他の地域地区(22)・都計法外の法令のチェック格子。
    # 法令格子は 37-1/38-1 で座標が異なる（38-1は生物多様性増進法を挿入）ため variant 別マップ。
    values.update(_nisshido_cells(variant, _g(h, "nisshido")))
    values.update(_horei_grid(_chiiki_chiku_marks(variant), _g(h, "chiiki_chiku")))
    values.update(_horei_grid(OTHER_HOREI_MARKS[variant], _g(h, "other_horei")))
    # 専有部分の建築時期（新築年月日）を 元号/年/月/日 に分割
    ck = _split_era_date(_g(se, "chikujiki"))
    if ck:
        values["L213"], values["O213"], values["S213"], values["W213"] = ck
    return values, clears_extra


# 変種 → 重説ビルダー
JUYOJIKO_BUILDERS = {
    "36-1": _build_juyojiko_36_1,
    "37-1": _build_juyojiko_kubun,
    "38-1": _build_juyojiko_kubun,
}


def build_aux(bc: Any) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    """補助シート（取引完了確認書・領収書）の当事者ヘッダーを差し込む。

    対象シートがワークブックに無ければ fill_workbook 側で無視される。
    bc は Juyojiko / Keiyakusho いずれでも可（urinushi/kainushi を参照）。
    """
    uri = _g(bc, "urinushi", "name")
    kai = _g(bc, "kainushi", "name")
    sv = {
        "335.取引完了確認書": {"G33": uri, "AB33": kai},
        "735-1.領収書": {"G21": kai},
    }
    sc = {s: list(v.keys()) for s, v in sv.items()}
    return sv, sc


# ── 区分様式のエディション判定 ──────────────────────────────
# 37-1/38-1 の重説には「指定建蔽率」が 388 行（A版）/ 390 行（B版）の2エディションが
# 流通する。現状の区分セルマップは A 版基準。B 版は行・列・マージが非一様に再編されて
# おり（区域区分チェックは +0、値セル・法令格子は +2、管理欄は列ごと別配置）、単純な
# 行オフセットでは正しく差し込めない。誤差込を避けるため、版を判定して呼び出し側で
# ガードする（B 版マップ確定までは A 版テンプレのみ自動差込を許可）。
def detect_kubun_edition(ws: Any) -> str:
    """区分重説シートの版を返す: 'A'(指定建蔽率388行) / 'B'(390行) / 'unknown'。"""
    for r in range(384, 401):
        for c in range(1, 24):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "指定建蔽率" in v:
                return {388: "A", 390: "B"}.get(r, "unknown")
    return "unknown"


# ── Edition B（37-1）の座標オーバーライド（実WB4物件で確証）──
# B版は一律行シフトではなくセクション別: チェック枠(区域区分/用途地域/防火/22条/高度)は
# A版と同位置、値セルは列移動、建築時期・法令格子・地域地区格子は +2 行。確証の取れた
# 範囲のみ写像し、未確証セル(取引条件の違約金/担保)はB版では出力しない(誤差込回避)。
KUBUN_B_VALUE_OVERRIDES = {
    "I194": "D194", "AL207": "U207", "L292": "D292",   # 一棟名称・専有名称・建物所有者住所
    "Q388": "D388", "Q402": "M402",                     # 指定建蔽率・容積率
    "L751": "L749",                                     # ペット飼育制限
    "L864": "B864", "U868": "I868", "V872": "U872",     # 修繕積立金 月額/累計/滞納
    "L884": "D884", "V888": "I888", "K900": "D900",     # 管理費月額/滞納・管理組合名称
    "L213": "L215", "O213": "O215", "S213": "S215", "W213": "W215",  # 建築時期 +2行
}
# 未確証のためB版では出力しないA版セル（Ⅱ取引条件の違約金/担保責任）
KUBUN_B_DROP = frozenset({"O1256", "W1256", "G1256", "AD1256", "T1328", "Z1328"})


def _kubun_b_remap(values: dict[str, Any],
                   clear_extra: list[str]) -> tuple[dict[str, Any], list[str]]:
    """区分A版の差込値をB版(37-1)座標へ写像する。法令/地域地区格子は+2行、値セルは列差替、
    未確証セルは出力しない。チェック枠・その他はA版と同位置のため不変。"""
    # 地域地区格子は専用フィールド管轄の先頭5ゾーン(防火/22条/高度=C372〜C380)を除く。
    # これらは BOKA/NIJUNI/KODO のチェック枠で A版と同位置のため +2 してはならない。
    grid_cells = set(OTHER_HOREI_MARKS.get("37-1", {}).values()) \
        | set(_chiiki_chiku_marks("37-1").values())

    def remap(coord: str) -> str | None:
        if coord in KUBUN_B_DROP:
            return None
        if coord in KUBUN_B_VALUE_OVERRIDES:
            return KUBUN_B_VALUE_OVERRIDES[coord]
        if coord in grid_cells:
            return _shift_row(coord, 2)
        return coord

    nv: dict[str, Any] = {}
    for c, v in values.items():
        b = remap(c)
        if b is not None:
            nv[b] = v
    nc = [b for c in clear_extra if (b := remap(c)) is not None]
    return nv, nc


def _shift_row(coord: str, delta: int) -> str:
    """セル座標の行を delta だけずらす（列は不変）。"""
    if not delta:
        return coord
    r, c = coordinate_to_tuple(coord)
    return f"{get_column_letter(c)}{r + delta}"


def build_juyojiko(variant: str, bc: Juyojiko,
                   edition: str = "A") -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    """(sheet_values, sheet_clear) を返す。wb_fill.fill_workbook にそのまま渡せる。

    edition: 区分(37-1)テンプレの様式版。'B' のときB版座標へ写像する（bc_service が
    detect_kubun_edition で判定して渡す）。'A'/その他は従来どおり。
    """
    builder = JUYOJIKO_BUILDERS.get(variant)
    if builder is None:
        raise KeyError(f"未対応のテンプレ変種: {variant}（対応: {list(JUYOJIKO_BUILDERS)}）")
    values, clear_extra = builder(bc, variant)
    if edition == "B" and variant == "37-1":
        values, clear_extra = _kubun_b_remap(values, clear_extra)
    clear = list(values.keys()) + clear_extra
    return {JUYOJIKO_SHEET: values}, {JUYOJIKO_SHEET: clear}
