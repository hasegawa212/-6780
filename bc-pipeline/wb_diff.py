"""セルマップ整備用の開発ツール: 同一テンプレの実例2通を差分して投入セルを特定する.

案件ごとに値が変わるセル＝データ投入セル。固定セル＝様式の文言。
新しいテンプレ/シートのセルマップを作る/更新するときに使う。

使い方::
    python wb_diff.py 例1.xlsx 例2.xlsx "不動産売買契約書"
"""

from __future__ import annotations

import sys

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter


def _label_for(ws, coord: str) -> str:
    r, c = coordinate_to_tuple(coord)
    for cc in range(c - 1, 0, -1):
        v = ws.cell(r, cc).value
        if isinstance(v, str):
            s = v.strip()
            if len(s) >= 2 and s not in ("□", "■", "・", "（", "）", "年", "月", "日", "㎡", "円"):
                return s[:20]
    return ""


def diff(path_a: str, path_b: str, sheet: str) -> list[tuple]:
    a = load_workbook(path_a, data_only=True)[sheet]
    b = load_workbook(path_b, data_only=True)[sheet]
    out = []
    mr = min(a.max_row, b.max_row)
    mc = min(a.max_column, b.max_column)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            va, vb = a.cell(r, c).value, b.cell(r, c).value
            if va != vb and (va not in (None, "") or vb not in (None, "")):
                co = f"{get_column_letter(c)}{r}"
                out.append((co, _label_for(a, co), str(va)[:24], str(vb)[:24]))
    return out


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    for row in diff(sys.argv[1], sys.argv[2], sys.argv[3]):
        print(row)
