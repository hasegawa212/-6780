"""差し込みエンジン.

白紙の BC テンプレ（openpyxl で開ける .xlsx）に、抽出 JSON ＋案件マスタの値を
流し込んで完成版の .xlsx バイト列を返す。テンプレが無い場合は
``make_blank_templates`` で自動生成する（実物テンプレが手に入ったら
``blank_36-1.xlsx`` 等を差し替えるだけでよい）。
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from bc_schema import (
    TEMPLATES,
    YOTO_OPTIONS,
    TemplateSpec,
    normalize_yoto,
    resolve_bukken,
)

HERE = Path(__file__).resolve().parent

CHECK_ON = "■"
CHECK_OFF = "□"


def _ensure_template(spec: TemplateSpec) -> Path:
    """テンプレの実体パスを返す。無ければスキーマから生成する."""
    path = HERE / spec.template_file
    if not path.exists():
        # 遅延 import（make_blank は openpyxl 書き込みのみで本番依存を増やさない）
        from make_blank_templates import build_template

        build_template(spec, path)
    return path


def _fmt(value: Any, suffix: str) -> str:
    """値を帳票表示用の文字列に整形する."""
    if isinstance(value, bool):
        text = "有" if value else "無"
    elif isinstance(value, (int, float)):
        # 12345678 → 12,345,678
        text = f"{value:,}"
    else:
        text = str(value)
    return f"{text}{suffix}" if text else text


def fill(
    bukken: str,
    extracted: dict[str, Any],
    deal_master: dict[str, Any] | None = None,
) -> tuple[bytes, int]:
    """BC を生成して (xlsx バイト列, 差し込んだ項目数) を返す."""
    key = resolve_bukken(bukken)
    spec = TEMPLATES[key]
    deal_master = deal_master or {}

    wb = load_workbook(_ensure_template(spec))
    ws = wb[spec.sheet] if spec.sheet in wb.sheetnames else wb.active

    sources = {"extracted": dict(extracted or {}), "deal_master": dict(deal_master)}
    # 用途地域は正式名称に寄せておく（チェックボックスと表示の両方で使う）
    if sources["extracted"].get("yoto"):
        sources["extracted"]["yoto"] = normalize_yoto(sources["extracted"]["yoto"])

    filled = 0
    for cell in spec.cells:
        value = sources.get(cell.source, {}).get(cell.key)
        if value is None or value == "":
            continue
        ws[cell.value_cell] = _fmt(value, cell.suffix)
        filled += 1

    # 用途地域チェックボックス（36-1: checkbox_361 / 37-1: checkbox_371）
    target_yoto = sources["extracted"].get("yoto")
    for i, opt in enumerate(YOTO_OPTIONS):
        row = spec.yoto_option_start_row + i
        mark = CHECK_ON if (target_yoto and opt == target_yoto) else CHECK_OFF
        ws[f"{spec.yoto_option_col}{row}"] = mark
        ws[f"{spec.yoto_label_col}{row}"] = opt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filled


def default_filename(bukken: str, extracted: dict[str, Any]) -> str:
    """所在地などから無難なファイル名を決める."""
    key = resolve_bukken(bukken)
    shozai = (extracted or {}).get("shozai") or "物件"
    # ファイル名に使えない文字をざっくり除去
    safe = "".join(c for c in str(shozai) if c not in r'\/:*?"<>|').strip()
    return f"BC_{key}_{safe[:40]}.xlsx"
