"""本番ワークブック（36-1/37-1/38-1 等）への差し込みエンジン.

御社の BC 用ワークブック（重説・契約書を含む多シート .xlsx）をテンプレートとして、
BC データを該当セルに差し込む。各テンプレの「データ投入セル」は cellmaps で定義し、
差し込み前にそのセルを **クリア** してから値を書くため、記入済みワークブックを
テンプレに使っても旧案件のデータは残らない（clear-then-fill）。

セル座標は同一テンプレの実例を差分比較して特定したもの（個人情報ではない）。
御社の実ワークブック自体はリポジトリに含めない。実行時に BC_TEMPLATE_DIR から読む。
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet


def set_print_fit_to_width(ws: Worksheet) -> None:
    """印刷設定を「全列を横1ページに収める」に揃える（右端の切れを防ぐ）。

    FRK 公式様式は固定縮小率（例 73%）だと A4 縦で右端（〜AX列）が切れることがある。
    fitToWidth=1 / fitToHeight=0（縦は必要なだけ）に統一し、そのまま印刷できるようにする。
    """
    ps = ws.page_setup
    ps.fitToWidth = 1
    ps.fitToHeight = 0
    ps.scale = None  # 固定縮尺と fitToPage は排他。縮尺指定を解除。
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(
        fitToPage=True, autoPageBreaks=True
    )


def _merged_anchor(ws: Worksheet, coord: str) -> str:
    """coord が結合セル範囲内なら左上アンカー座標を返す（openpyxl は書込先がアンカー）。"""
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return rng.coord.split(":")[0]
    return coord


def fill_sheet(
    ws: Worksheet,
    values: dict[str, Any],
    clear_coords: list[str],
) -> int:
    """1 シートに clear-then-fill する。書き込んだ件数を返す。"""
    # まず全データセルをクリア（旧案件の値を消す）
    for coord in clear_coords:
        anchor = _merged_anchor(ws, coord)
        ws[anchor] = None
    # BC 値を書き込み
    written = 0
    for coord, val in values.items():
        if val is None or val == "":
            continue
        anchor = _merged_anchor(ws, coord)
        ws[anchor] = val
        written += 1
    return written


def fill_workbook(
    template: bytes | str | Path,
    sheet_values: dict[str, dict[str, Any]],
    sheet_clear: dict[str, list[str]],
    fit_to_width: bool = True,
) -> tuple[bytes, int]:
    """テンプレートに複数シート分を差し込み、(xlsx バイト列, 書込総数) を返す。

    template: .xlsx のバイト列 or パス。
    sheet_values: {シート名: {coord: value}}
    sheet_clear:  {シート名: [クリアする coord]}
    fit_to_width: 差し込んだ各シートの印刷設定を横1ページ収めに統一（既定 True）。
    """
    src = io.BytesIO(template) if isinstance(template, (bytes, bytearray)) else template
    wb = load_workbook(src)
    total = 0
    for sheet, values in sheet_values.items():
        if sheet not in wb.sheetnames:
            continue
        total += fill_sheet(wb[sheet], values, sheet_clear.get(sheet, []))
        if fit_to_width:
            set_print_fit_to_width(wb[sheet])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), total


def load_template(template_dir: str | Path, variant: str) -> bytes | None:
    """BC_TEMPLATE_DIR/<variant>.xlsx を読む。無ければ None。"""
    path = Path(template_dir) / f"{variant}.xlsx"
    if not path.exists():
        return None
    return path.read_bytes()


# 重説/契約書シートの A1 マーカー先頭 → 変種キー
_VARIANT_MARKERS = {"36-1": "36-1", "37-1": "37-1", "38-1": "38-1", "16-1": "16-1"}


def detect_variant(template: bytes) -> str | None:
    """ワークブックの「重要事項説明書」シート A1 から様式（36-1/37-1/38-1）を判定する。

    例: A1='37-1.売主宅建業者用/区分所有建物（敷地権）' → '37-1'。
    判定できなければ None（呼び出し側で明示指定にフォールバック）。
    """
    wb = load_workbook(io.BytesIO(template), read_only=True, data_only=True)
    try:
        ws = wb["重要事項説明書"] if "重要事項説明書" in wb.sheetnames else wb[wb.sheetnames[0]]
        a1 = ws["A1"].value
    finally:
        wb.close()
    if not isinstance(a1, str):
        return None
    head = a1.strip()
    for key in _VARIANT_MARKERS:
        if head.startswith(key):
            return key
    return None

