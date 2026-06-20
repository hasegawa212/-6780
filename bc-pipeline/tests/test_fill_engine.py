"""fill_engine / bc_schema の最小テスト（pytest 不要・素の assert）.

実行::
    cd bc-pipeline && python tests/test_fill_engine.py

抽出値は実物の AB 側重要事項説明書（フレクション長岡 503 号室）の
物件属性に基づく（個人情報は含めない）。
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

import fill_engine  # noqa: E402
from bc_schema import normalize_yoto, resolve_bukken  # noqa: E402

# 実物重説（区分所有）から読み取れる物件属性（page4-5）
NAGAOKA = {
    "shozai": "新潟県長岡市曲新町551-1",
    "kuiki": "市街化区域",           # ①区域区分 ■1.市街化区域
    "yoto": "第1種住居地域",         # ①用途地域 ■5.第1種住居地域
    "nijuni_jo": True,               # ②4.建築基準法第22条区域 ■
    "kenpei": 60,                    # ③指定建蔽率 60%
    "yoseki": 200,                   # ④指定容積率 200%
}
# BC（B→C 転売）側の確定値。案件マスタから来る想定の例。
DEAL = {"buyer_C": "東洋建設ホーム株式会社", "bc_baibai_daikin": 27800000}


def test_resolve_bukken() -> None:
    assert resolve_bukken("戸建") == "戸建"
    assert resolve_bukken("マンション") == "区分"
    assert resolve_bukken("37-1") == "区分"


def test_normalize_yoto() -> None:
    assert normalize_yoto("第1種住居") == "第1種住居地域"
    assert normalize_yoto("第1種中高層") == "第1種中高層住居専用地域"
    assert normalize_yoto("1低") == "第1種低層住居専用地域"
    assert normalize_yoto("商業地域") == "商業地域"
    assert normalize_yoto("指定なし") == "用途地域の指定なし"


def _flat_values(xlsx: bytes) -> list[object]:
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


def test_generate_kubun_autocreates_template() -> None:
    # blank_37-1.xlsx が無くてもスキーマから生成できること（残タスク対応）
    xlsx, filled = fill_engine.fill("区分", NAGAOKA, DEAL)
    assert filled == 8, f"差し込み項目数が想定外: {filled}"
    assert (ROOT / "blank_37-1.xlsx").exists()
    flat = _flat_values(xlsx)
    assert "新潟県長岡市曲新町551-1" in flat
    assert "60%" in flat and "200%" in flat
    assert "27,800,000 円" in flat
    # 用途地域チェック: 正式名称に ■ が立つ
    assert "■" in flat
    assert "第1種住居地域" in flat


def test_generate_kodate() -> None:
    xlsx, filled = fill_engine.fill("戸建", NAGAOKA, DEAL)
    assert filled == 8
    assert (ROOT / "blank_36-1.xlsx").exists()


def test_default_filename() -> None:
    name = fill_engine.default_filename("区分", NAGAOKA)
    assert name.startswith("BC_区分_") and name.endswith(".xlsx")


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"PASS {fn.__name__}")
    print(f"\n{len(fns)} tests passed")
