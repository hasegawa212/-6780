"""BC重説パイプラインの最小テスト（pytest 不要・素の assert）.

実行::
    cd bc-pipeline && python tests/test_pipeline.py

AB 側データは実物の重要事項説明書（フレクション長岡 503 号室・区分所有）の
物件属性に基づく（個人情報は含めない）。
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

import juyojiko_excel  # noqa: E402
import keiyaku_excel  # noqa: E402
from bc_schema import normalize_yoto, resolve_bukken  # noqa: E402
from bc_transform import transform_ab_to_bc, transform_keiyaku_ab_to_bc  # noqa: E402
from keiyaku_schema import KeiyakuDaikin, Keiyakusho  # noqa: E402
from juyojiko_schema import TochiHyoji  # noqa: E402
from juyojiko_schema import (  # noqa: E402
    FudosanHyoji,
    HoreiSeigen,
    Juyojiko,
    KanriHiyou,
    Party,
    TatemonoHyoji,
    TorihikiJoken,
)

# 実物 AB 重説（区分）の物件属性（page3-12）
AB = Juyojiko(
    bukken_type="区分",
    torihiki_taiyo="売買 ・ 媒介",
    urinushi=Party(name="有限会社ネットプラン", address="岐阜県多治見市太平町一丁目47番地の2"),
    kainushi=Party(name="株式会社Martial Arts"),
    fudosan=FudosanHyoji(
        bukken_type="区分",
        jukyo_hyoji="新潟県長岡市曲新町551-1",
        ittou_shozai="長岡市曲新町字横田 551番地1",
        ittou_kozo="鉄筋コンクリート造陸屋根5階建",
        senyuu=TatemonoHyoji(kaoku_bango="曲新町 551番1の503", meisho="503号",
                             shurui="居宅", yukamenseki="14.95㎡"),
    ),
    horei=HoreiSeigen(
        toshikeikaku_kuiki="都市計画区域内",
        kuiki_kubun="市街化区域",
        yoto="第1種住居地域",
        nijuni_jo=True,
        kenpei=60,
        yoseki=200,
    ),
    kanri=KanriHiyou(kanrihi_getsugaku=5000, shuzen_getsugaku=3000,
                     kanri_kumiai="フレクション長岡管理組合", kanri_keitai="全部委託管理",
                     kanri_itakusaki="日本ハウズイング株式会社"),
    joken=TorihikiJoken(baibai_daikin=1_500_000),
    tokuyaku=["本物件は賃借権の負担付きで売買。買主は貸主の地位を承継する。"],
)

# 案件マスタ（BC 側）
DEAL = {
    "buyer_C": "東洋建設ホーム株式会社",
    "buyer_C_address": "東京都〇〇区〇〇1-1-1",
    "bc_baibai_daikin": 27_800_000,
    "bc_tetsuke": 1_000_000,
}


def _flat(xlsx: bytes) -> list[object]:
    ws = load_workbook(io.BytesIO(xlsx)).active
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


def test_resolve_bukken() -> None:
    assert resolve_bukken("マンション") == "区分"
    assert resolve_bukken("土地建物") == "戸建"


def test_normalize_yoto() -> None:
    assert normalize_yoto("第1種住居") == "第1種住居地域"
    assert normalize_yoto("指定なし") == "用途地域の指定なし"


def test_transform_swaps_parties_and_price() -> None:
    bc = transform_ab_to_bc(AB, DEAL)
    # 売主 A→B、買主 B→C
    assert bc.urinushi.name == "株式会社Martial Arts"
    assert bc.kainushi.name == "東洋建設ホーム株式会社"
    # 代金は BC 価格
    assert bc.joken.baibai_daikin == 27_800_000
    assert bc.joken.tetsuke == 1_000_000
    # AB は不変（元データを壊さない）
    assert AB.urinushi.name == "有限会社ネットプラン"
    assert AB.joken.baibai_daikin == 1_500_000


def test_transform_carries_property_facts() -> None:
    bc = transform_ab_to_bc(AB, DEAL)
    # 物件事実はそのまま引き継ぐ
    assert bc.horei.kuiki_kubun == "市街化区域"
    assert bc.horei.yoto == "第1種住居地域"
    assert bc.horei.kenpei == 60 and bc.horei.yoseki == 200
    assert bc.kanri.kanrihi_getsugaku == 5000
    assert bc.fudosan.ittou_shozai == "長岡市曲新町字横田 551番地1"
    # 三為注記が付く
    assert any("所有権移転先" in t or "中間省略" in t for t in bc.tokuyaku)


def test_transform_carries_shakuchi_facts() -> None:
    from juyojiko_schema import Shakuchi
    ab = AB.model_copy(deep=True)
    ab.shakuchi = Shakuchi(
        shakuchiken_shurui="普通借地権",
        sonzoku_kikan="令和3年4月1日〜令和23年3月31日",
        jidai_kingaku=30_000,
        jidai_tani="月額",
        koshin_ryo="更新時に別途協議",
        teichi_shoyusha_shimei="地主 太郎",
    )
    bc = transform_ab_to_bc(ab, DEAL)
    # 借地条件は物件事実としてそのまま引き継ぐ（当事者・代金のみ差替）
    assert bc.shakuchi.shakuchiken_shurui == "普通借地権"
    assert bc.shakuchi.jidai_kingaku == 30_000
    assert bc.shakuchi.teichi_shoyusha_shimei == "地主 太郎"
    # 所有権物件（shakuchi 無し）では None のまま
    assert transform_ab_to_bc(AB, DEAL).shakuchi is None


def test_render_bc_excel_shakuchi_section() -> None:
    from juyojiko_schema import Shakuchi
    ab = AB.model_copy(deep=True)
    ab.shakuchi = Shakuchi(shakuchiken_shurui="普通借地権", jidai_kingaku=30_000,
                           jidai_tani="月額", koshin_ryo="更新時に協議",
                           teichi_shoyusha_shimei="地主 太郎")
    bc = transform_ab_to_bc(ab, DEAL)
    flat = _flat(juyojiko_excel.render(bc))
    assert "借地権の内容（借地借家法）" in flat
    assert "普通借地権" in flat
    assert any(isinstance(v, str) and "30,000 円" in v for v in flat)
    # 所有権物件では借地セクションは出ない
    assert "借地権の内容（借地借家法）" not in _flat(juyojiko_excel.render(transform_ab_to_bc(AB, DEAL)))


def test_demo_runs_offline(tmp_path) -> None:
    import demo
    bc_j = transform_ab_to_bc(demo.sample_ab_juyojiko(shakuchi=True), demo.sample_deal())
    bc_k = transform_keiyaku_ab_to_bc(demo.sample_ab_keiyaku(), demo.sample_deal())
    # 当事者A→B→C・代金差替が効いている
    assert bc_j.urinushi.name == "株式会社Martial Arts"
    assert bc_j.kainushi.name == "東洋建設ホーム株式会社"
    assert bc_j.joken.baibai_daikin == 27_800_000
    assert bc_k.daikin.baibai_daikin == 27_800_000
    # 借地条件が引き継がれている
    assert bc_j.shakuchi.shakuchiken_shurui == "普通借地権"
    # 個人情報を含まない（サンプルはダミー社名のみ）
    assert "様" not in (bc_j.kainushi.name or "")


def test_demo_live_requires_key(tmp_path) -> None:
    import os
    import demo
    if os.environ.get("ANTHROPIC_API_KEY"):
        return  # 鍵がある環境では実呼び出しになるのでスキップ
    pdf = tmp_path / "x.pdf"
    pdf.write_bytes(b"%PDF-1.4\n%%EOF")
    # 鍵未設定では抽出は空になり、_live_extract は警告文で例外化する
    # （/extract 自体は500を投げない設計。呼び出し側 main が握ってサンプルに退避する）
    try:
        demo._live_extract(str(pdf), "juyojiko")
        assert False, "鍵無しで成功するのはおかしい"
    except Exception as e:
        assert "APIキー" in str(e) or "手入力" in str(e)


def test_wb_probe_finds_shakuchi_labels(tmp_path) -> None:
    import wb_probe
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "171-1.借地説明書"
    ws["B3"] = "借地権の種類"; ws["F3"] = "普通借地権"; ws["E3"] = "□"
    ws["H3"] = "定期借地権"; ws["G3"] = "□"
    ws["B7"] = "地代"; ws["F7"] = None
    p = tmp_path / "shakuchi.xlsx"; wb.save(p)
    rows = wb_probe.probe(str(p), "171-1.借地説明書", wb_probe.PRESETS["shakuchi"])
    labels = {label for _, label, _ in rows}
    assert "借地権の種類" in labels and "地代" in labels
    # 借地権の種類 行の候補に近傍チェック枠（□）が含まれる
    kinds = next(c for co, label, c in rows if label == "借地権の種類")
    assert "E3" in kinds or "G3" in kinds
    # プリセット外の語（preset未指定で全走査）も拾える
    assert wb_probe.probe(str(p), "171-1.借地説明書", None)


def test_juyojiko_kubun_clears_occupant_pii_cells() -> None:
    # 区分の占有者 住所(F277)・氏名(F279)は個別セル。他号室テンプレ使用時に
    # 前入居者のPIIが残らないよう、常にクリア対象に含めること。
    import cellmaps
    ab = Juyojiko(bukken_type="区分", kainushi=Party(name="株式会社Martial Arts"),
                  fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x",
                                       senyuu=TatemonoHyoji()),
                  horei=HoreiSeigen(kenpei=80, yoseki=400),
                  senyuusha_uchi="第三者占有 有：占有者の概要")
    bc = transform_ab_to_bc(ab, DEAL)
    for variant in ("37-1", "38-1"):
        _, clears = cellmaps.build_juyojiko(variant, bc)
        sheet_clears = clears["重要事項説明書"]
        assert "F277" in sheet_clears and "F279" in sheet_clears, variant
        # 管理(組合名称K900/委託先K904)・委託先登録番号(AQ908)・添付書類(E1425)も
        # 他物件テンプレ流用時にデータが残らないようクリア対象に含める。
        assert "K900" in sheet_clears and "K904" in sheet_clears, variant
        assert "AQ908" in sheet_clears and "E1425" in sheet_clears, variant


def test_detect_variant_from_a1() -> None:
    import io
    import wb_fill
    from openpyxl import Workbook
    for marker, expect in [
        ("37-1.売主宅建業者用/区分所有建物（敷地権）", "37-1"),
        ("36-1.売主宅建業者用/土地建物/売買代金固定", "36-1"),
        ("38-1.売主宅建業者用/区分所有建物（非敷地権）", "38-1"),
    ]:
        wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"; ws["A1"] = marker
        buf = io.BytesIO(); wb.save(buf)
        assert wb_fill.detect_variant(buf.getvalue()) == expect
    # マーカー無しは None（明示指定にフォールバック）
    wb = Workbook(); wb.active["A1"] = "見出し"; buf = io.BytesIO(); wb.save(buf)
    assert wb_fill.detect_variant(buf.getvalue()) is None


def test_juyojiko_newly_mapped_fields() -> None:
    # 都市計画区域内/外・違約金%・担保責任の措置・区分の建築時期 を新規マップ
    import cellmaps
    from juyojiko_schema import TatemonoHyoji
    ab = Juyojiko(
        bukken_type="区分", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x",
                             senyuu=TatemonoHyoji(chikujiki="平成2年12月3日新築")),
        horei=HoreiSeigen(toshikeikaku_kuiki="都市計画区域内", kuiki_kubun="市街化区域",
                          kenpei=60, yoseki=200),
        joken=TorihikiJoken(iyakukin_wariai=20, tanpo_sekinin="講じない"))
    v, _ = cellmaps.build_juyojiko("37-1", ab)
    vv = v["重要事項説明書"]
    assert vv["J335"] == "■" and vv["J337"] == "□"        # 都市計画区域内
    assert vv["O1256"] == "■" and vv["W1256"] == 20         # 違約金 売買代金の20%
    assert vv["T1328"] == "□" and vv["Z1328"] == "■"        # 担保 講じない
    assert (vv["L213"], vv["O213"], vv["S213"], vv["W213"]) == ("平成", 2, 12, 3)  # 建築時期

    ab36 = Juyojiko(
        bukken_type="戸建", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
        horei=HoreiSeigen(toshikeikaku_kuiki="都市計画区域外", kenpei=50, yoseki=100),
        joken=TorihikiJoken(iyakukin_wariai=10, tanpo_sekinin="講じる"))
    w, _ = cellmaps.build_juyojiko("36-1", ab36)
    ww = w["重要事項説明書"]
    assert ww["J331"] == "□" and ww["J333"] == "■"          # 都市計画区域外
    assert ww["O1008"] == "■" and ww["W1008"] == 10
    assert ww["T1078"] == "■" and ww["Z1078"] == "□"        # 担保 講じる


def test_juyojiko_batch2_value_fields() -> None:
    # 地積(実測)・建築時期(戸建)・日影 有/無 を新規マップ（36-1）
    import cellmaps
    bc = Juyojiko(
        bukken_type="戸建", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建",
                             tochi=TochiHyoji(shozai="x", chiseki_jissoku="123.45"),
                             tatemono=TatemonoHyoji(chikujiki="令和3年5月1日")),
        horei=HoreiSeigen(nisshido="第一種 4h-2.5h"))
    vv = cellmaps.build_juyojiko("36-1", bc)[0]["重要事項説明書"]
    assert vv["G206"] == "123.45"                               # 実測面積
    assert (vv["H246"], vv["K246"], vv["O246"], vv["S246"]) == ("令和", 3, 5, 1)
    assert vv["L412"] == "■" and vv["O412"] == "□"             # 日影 有
    # 日影 無（無し/対象外を含む文字列）
    bc2 = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                   fudosan=FudosanHyoji(bukken_type="戸建"),
                   horei=HoreiSeigen(nisshido="規制無し"))
    vv2 = cellmaps.build_juyojiko("36-1", bc2)[0]["重要事項説明書"]
    assert vv2["L412"] == "□" and vv2["O412"] == "■"           # 日影 無
    # 日影 区分(37-1/38-1)は別行(416)
    bck = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                   fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x"),
                   horei=HoreiSeigen(nisshido="有"))
    vk = cellmaps.build_juyojiko("38-1", bck)[0]["重要事項説明書"]
    assert vk["L416"] == "■" and vk["O416"] == "□"


def test_juyojiko_36_1_torihiki_daikin() -> None:
    # Ⅱ-1 売買代金・土地/建物価格・消費税・手付金（36-1）。未指定でも残値防止でクリア対象。
    import cellmaps
    bc = Juyojiko(
        bukken_type="戸建", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
        horei=HoreiSeigen(kenpei=50, yoseki=100),
        joken=TorihikiJoken(baibai_daikin=19_800_000, tochi_kakaku=7_800_000,
                            tatemono_kakaku=12_000_000, shohizei=1_090_909, tetsuke=100_000))
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    vv = sv["重要事項説明書"]
    assert vv["H868"] == 19_800_000                         # 売買代金
    assert vv["AL866"] == 7_800_000 and vv["AL868"] == 12_000_000  # 土地/建物価格
    assert vv["AL870"] == 1_090_909 and vv["V881"] == 100_000      # 消費税/手付金
    # 値が無くてもキーは出る（=clearに含まれ、他物件テンプレ残値を消す）
    assert "H868" in sc["重要事項説明書"] and "AL868" in sc["重要事項説明書"]


def test_juyojiko_horei_check_grids() -> None:
    # その他の地域地区(22)・都計法外の法令(61)のチェック格子
    import cellmaps
    import cellmap_grids
    bc = Juyojiko(
        bukken_type="戸建", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建"),
        # 防火/22条/高度は専用フィールド側の管轄のため、格子テストは「その他」ゾーンを使う
        horei=HoreiSeigen(chiiki_chiku=["高度利用地区", "景観地区"],
                          other_horei=["古都保存法", "盛土規制法", "文化財保護法"]))
    vv = cellmaps.build_juyojiko("36-1", bc)[0]["重要事項説明書"]
    # 選択した地区・法令は ■
    assert vv[cellmap_grids.CHIIKI_CHIKU_MARKS["36-1"]["高度利用地区"]] == "■"
    assert vv[cellmap_grids.CHIIKI_CHIKU_MARKS["36-1"]["景観地区"]] == "■"
    assert vv[cellmap_grids.OTHER_HOREI_MARKS["36-1"]["古都保存法"]] == "■"
    assert vv[cellmap_grids.OTHER_HOREI_MARKS["36-1"]["文化財保護法"]] == "■"
    # 略称（盛土規制法）も正規化して命中
    assert vv[cellmap_grids.OTHER_HOREI_MARKS["36-1"]["宅地造成及び特定盛土等規制法"]] == "■"
    # 未選択枠は □（残留防止のため格子全体を初期化）
    assert vv[cellmap_grids.CHIIKI_CHIKU_MARKS["36-1"]["臨港地区"]] == "□"
    assert vv[cellmap_grids.OTHER_HOREI_MARKS["36-1"]["農地法"]] == "□"

    # データ無し（空リスト）のときは格子に触れない
    empty = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                     fudosan=FudosanHyoji(bukken_type="戸建"), horei=HoreiSeigen())
    ev = cellmaps.build_juyojiko("36-1", empty)[0]["重要事項説明書"]
    assert cellmap_grids.OTHER_HOREI_MARKS["36-1"]["農地法"] not in ev

    # 37-1 と 38-1 で同一法令が異なる座標に入る（38-1は生物多様性増進法を挿入）
    bck = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                   fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x"),
                   horei=HoreiSeigen(other_horei=["文化財保護法", "生物多様性増進法"]))
    b37 = cellmaps.build_juyojiko("37-1", bck)[0]["重要事項説明書"]
    b38 = cellmaps.build_juyojiko("38-1", bck)[0]["重要事項説明書"]
    c37 = cellmap_grids.OTHER_HOREI_MARKS["37-1"]["文化財保護法"]
    c38 = cellmap_grids.OTHER_HOREI_MARKS["38-1"]["文化財保護法"]
    assert c37 != c38 and b37[c37] == "■" and b38[c38] == "■"
    # 37-1 様式に無い生物多様性増進法は例外無くスキップ（38-1 のみ枠あり）
    assert "地域における生物の多様性の増進のための活動の促進等に関する法律" \
        not in cellmap_grids.OTHER_HOREI_MARKS["37-1"]
    assert b38[cellmap_grids.OTHER_HOREI_MARKS["38-1"][
        "地域における生物の多様性の増進のための活動の促進等に関する法律"]] == "■"


def test_render_bc_excel() -> None:
    bc = transform_ab_to_bc(AB, DEAL)
    flat = _flat(juyojiko_excel.render(bc))
    assert "重 要 事 項 説 明 書" in flat
    assert "株式会社Martial Arts" in flat       # 売主B
    assert "東洋建設ホーム株式会社" in flat       # 買主C
    assert "27,800,000 円" in flat               # BC代金
    assert "新潟県長岡市曲新町551-1" in flat       # 物件
    # 用途地域チェックが正式名称に ■
    assert any(isinstance(v, str) and "■ 第1種住居地域" in v for v in flat)
    # 区域区分チェック
    assert any(isinstance(v, str) and "■ 市街化区域" in v for v in flat)


def test_chiiki_chiku_grid_does_not_clobber_dedicated_marks() -> None:
    # その他地域地区の格子は、防火/22条/高度の専用チェック(C368〜C376)を打ち消さないこと。
    # （chiiki_chiku データがあると格子が□初期化し、nijuni_jo等を潰すバグの回帰防止）
    import cellmaps
    bc = Juyojiko(
        bukken_type="戸建", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
        horei=HoreiSeigen(nijuni_jo=True, boka="防火地域", kodo_chiku="第1種高度",
                          chiiki_chiku=["特定用途誘導地区", "都市再生特別地区"]))
    v = cellmaps.build_juyojiko("36-1", bc)[0]["重要事項説明書"]
    assert v["C368"] == "■"   # 防火（boka）
    assert v["C374"] == "■"   # 建築基準法第22条区域（nijuni_jo）← 修正対象
    assert v["C376"] == "■"   # 高度地区（kodo_chiku）
    assert v["AG378"] == "■"  # その他: 特定用途誘導地区（格子）
    assert v["R368"] == "□"   # その他: 風致地区（未選択）


# 実物 AB 売買契約書（戸建・ひたちなか市）の属性（個人情報は含めない）
AB_KEIYAKU = Keiyakusho(
    bukken_type="戸建",
    urinushi=Party(name="売主A"),
    kainushi=Party(name="株式会社Martial Arts"),
    fudosan=FudosanHyoji(
        bukken_type="戸建",
        tochi=TochiHyoji(shozai="ひたちなか市大字勝田字寺漏 3317番11", chimoku="宅地",
                         chiseki_toki="213.96㎡"),
        tatemono=TatemonoHyoji(shurui="居宅", yukamenseki="122.97㎡"),
    ),
    daikin=KeiyakuDaikin(baibai_daikin=16_900_000, tetsuke=1_900_000,
                         zankin=15_000_000, zankin_date="2025年4月10日"),
    tokuyaku=["別添「設備表」において有とした設備を含む。"],
)


def test_keiyaku_transform_and_render() -> None:
    bc = transform_keiyaku_ab_to_bc(AB_KEIYAKU, DEAL)
    assert bc.urinushi.name == "株式会社Martial Arts"      # 売主A→B
    assert bc.kainushi.name == "東洋建設ホーム株式会社"      # 買主B→C
    assert bc.daikin.baibai_daikin == 27_800_000           # BC代金
    # 価格が変わったので残代金を再計算（古い AB 残代金 15,000,000 を上書き）
    assert bc.daikin.zankin == 27_800_000 - 1_000_000
    # 物件・約款は引き継ぐ／AB不変
    assert bc.fudosan.tochi.chiseki_toki == "213.96㎡"
    assert AB_KEIYAKU.daikin.baibai_daikin == 16_900_000
    flat = _flat(keiyaku_excel.render(bc))
    assert "不 動 産 売 買 契 約 書" in flat
    assert "東洋建設ホーム株式会社" in flat
    assert "27,800,000 円" in flat
    # 約款が無くても標準条文見出し骨子が出る
    assert any(isinstance(v, str) and v.startswith("第1条") for v in flat)
    # 契約書の特約欄は御社定型「重要事項説明書に準拠する」へ集約（三為本文は重説側）
    assert any("重要事項説明書に準拠" in t for t in bc.tokuyaku)


def test_keiyaku_iyakukin() -> None:
    # 契約書 表紙「違約金の額」= 売買代金の N%（重説 Ⅱ取引条件と整合）
    import cellmaps
    bc = AB_KEIYAKU.model_copy(deep=True)
    bc.daikin.iyakukin_wariai = 20
    v36 = cellmaps.build_keiyaku("36-1", bc)[0]["不動産売買契約書"]
    assert v36["X65"] == "■" and v36["AF65"] == 20            # 2.売買代金の20%
    assert v36["P65"] == "□" and v36["AM65"] == "□"           # 他の選択肢は□

    # 区分（37-1/38-1 共通レイアウト・別行）
    bck = Keiyakusho(
        bukken_type="区分", urinushi=Party(name="売主"), kainushi=Party(name="買主"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x"),
        daikin=KeiyakuDaikin(baibai_daikin=30_000_000, iyakukin_wariai=15))
    for variant in ("37-1", "38-1"):
        vk = cellmaps.build_keiyaku(variant, bck)[0]["不動産売買契約書"]
        assert vk["X70"] == "■" and vk["AF70"] == 15
        assert vk["P70"] == "□" and vk["AM70"] == "□"

    # 違約金未指定のときは表紙の選択を非改変（テンプレ既定の20%を温存）
    no_iw = cellmaps.build_keiyaku("36-1", AB_KEIYAKU)[0]["不動産売買契約書"]
    assert "X65" not in no_iw and "AF65" not in no_iw


def test_keiyaku_omote_full_coverage() -> None:
    # 表紙の追加欄（内金①・引渡日・公租公課起算日・融資・業者/取引士・締結日）
    import cellmaps
    from juyojiko_schema import Gyosha, Torikiishi
    common = dict(
        gyosha=Gyosha(shozai="東京都中野区江原町3-34-1", shomei="株式会社Martial Arts",
                      daihyo="長谷川 光"),
        torikiishi=Torikiishi(shimei="小玉 浩之"),
        hikiwatashi_date="令和7年9月30日",
        seisan_kisanbi="令和7年1月1日",
        keiyaku_date="令和7年6月25日",
        loan_tokuyaku=True, loan_kingaku=27_300_000,
        loan_kaijo_date="令和7年11月25日",
    )
    bc36 = AB_KEIYAKU.model_copy(deep=True)
    bc36.daikin.uchikin1 = 2_000_000
    bc36.daikin.uchikin1_date = "令和7年8月1日"
    for k, v in common.items():
        setattr(bc36, k, v)
    v = cellmaps.build_keiyaku("36-1", bc36)[0]["不動産売買契約書"]
    assert v["AE55"] == 2_000_000 and (v["S55"], v["W55"], v["AA55"]) == (7, 8, 1)  # 内金①
    assert v["AE61"] == "令和7年9月30日"                       # 引渡日
    assert (v["S63"], v["W63"], v["AA63"]) == (7, 1, 1)        # 公租公課起算日
    assert v["Q67"] == "■" and v["U67"] == "□"                 # 融資利用 有
    assert v["AE71"] == 27_300_000                             # 融資金額
    assert (v["AH81"], v["AL81"], v["AP81"]) == (7, 11, 25)    # 融資解除期日
    assert v["P135"] == "東京都中野区江原町3-34-1"             # 業者所在地
    assert v["P137"] == "株式会社Martial Arts"                 # 商号
    assert v["P139"] == "長谷川 光" and v["P143"] == "小玉 浩之"  # 代表者/取引士
    assert (v["AI130"], v["AL130"], v["AP130"], v["AT130"]) == ("令和", 7, 6, 25)  # 締結日

    # 区分（37-1/38-1 共通レイアウト・別行）。引渡日は既存 AE66 と同一セル。
    bck = Keiyakusho(
        bukken_type="区分", urinushi=Party(name="売主"), kainushi=Party(name="買主"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x"),
        daikin=KeiyakuDaikin(baibai_daikin=30_000_000), **common)
    for variant in ("37-1", "38-1"):
        vk = cellmaps.build_keiyaku(variant, bck)[0]["不動産売買契約書"]
        assert vk["AE66"] == "令和7年9月30日"                  # 引渡日
        assert (vk["S68"], vk["W68"], vk["AA68"]) == (7, 1, 1)  # 公租公課起算日
        assert vk["Q72"] == "■" and vk["U72"] == "□"           # 融資 有
        assert vk["P142"] == "株式会社Martial Arts" and vk["P148"] == "小玉 浩之"
        assert (vk["AI135"], vk["AL135"], vk["AP135"], vk["AT135"]) == ("令和", 7, 6, 25)

    # 未充当（空 Keiyakusho）でも表紙セルはクリア対象に入る（他物件残留防止）
    _, clears = cellmaps.build_keiyaku("36-1", AB_KEIYAKU)
    sc = clears["不動産売買契約書"]
    assert "S63" in sc and "AH81" in sc and "AI130" in sc and "P137" in sc


def test_workbook_fill_clear_then_fill() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    # 合成テンプレ（不動産売買契約書シートに旧案件データを置く）
    wb = Workbook()
    ws = wb.active
    ws.title = "不動産売買契約書"
    ws["E123"] = "旧売主"
    ws["AB123"] = "旧買主"
    ws["AE45"] = 23300000
    ws["X11"] = "旧地番12"      # クリア専用セル
    buf = io.BytesIO()
    wb.save(buf)

    bc = transform_keiyaku_ab_to_bc(AB_KEIYAKU, DEAL)
    sv, sc = cellmaps.build_keiyaku("36-1", bc)
    out, n = wb_fill.fill_workbook(buf.getvalue(), sv, sc)

    ws2 = load_workbook(io.BytesIO(out))["不動産売買契約書"]
    assert ws2["E123"].value == "株式会社Martial Arts"   # 売主B
    assert ws2["AB123"].value == "東洋建設ホーム株式会社"  # 買主C
    assert ws2["AE45"].value == 27_800_000                # BC代金
    # 土地所在 "…3317番11" は 所在/番/番地 に分割される
    assert ws2["X11"].value == "3317" and ws2["AC11"].value == "11"
    assert n >= 5


def test_workbook_fill_juyojiko_36_1() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "重要事項説明書"
    ws["F7"] = "旧買主"
    ws["F263"] = "旧売主"
    ws["Q384"] = 50
    ws["X194"] = "旧地番"     # クリア専用
    buf = io.BytesIO()
    wb.save(buf)

    ab = Juyojiko(
        bukken_type="戸建",
        kainushi=Party(name="株式会社Martial Arts"),
        fudosan=FudosanHyoji(bukken_type="戸建",
                             tochi=TochiHyoji(shozai="テスト町1-2", chiseki_toki="200㎡")),
        horei=HoreiSeigen(kenpei=60, yoseki=200),
    )
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, n = wb_fill.fill_workbook(buf.getvalue(), sv, sc)

    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["F7"].value == "東洋建設ホーム株式会社"     # 買主C
    assert ws2["F263"].value == "株式会社Martial Arts"     # 売主B
    assert ws2["Q384"].value == 60 and ws2["Q398"].value == 200
    assert ws2["X194"].value is None
    assert n >= 5


def test_workbook_fill_keiyaku_kubun_37_1() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import ShikichikenTochi

    wb = Workbook()
    ws = wb.active
    ws.title = "不動産売買契約書"
    ws["E128"] = "旧売主"
    ws["AB128"] = "旧買主"
    ws["AE56"] = 12345          # 旧消費税（クリア専用）
    buf = io.BytesIO()
    wb.save(buf)

    ab = Keiyakusho(
        bukken_type="区分",
        urinushi=Party(name="元A"), kainushi=Party(name="株式会社Martial Arts"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="テスト市1番",
                             senyuu=TatemonoHyoji(kaoku_bango="1番の101", meisho="101"),
                             shikichiken=[ShikichikenTochi(chiseki="500.00", wariai="1/100")]),
        daikin=KeiyakuDaikin(baibai_daikin=1_700_000, tetsuke=100_000, zankin=1_600_000),
    )
    for variant in ("37-1", "38-1"):
        bc = transform_keiyaku_ab_to_bc(ab, DEAL)
        sv, sc = cellmaps.build_keiyaku(variant, bc)
        out, n = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
        ws2 = load_workbook(io.BytesIO(out))["不動産売買契約書"]
        assert ws2["E128"].value == "株式会社Martial Arts", variant
        assert ws2["AB128"].value == "東洋建設ホーム株式会社", variant
        assert ws2["AE54"].value == 27_800_000, variant
        assert ws2["AE56"].value is None, variant   # 旧消費税はクリア
        assert n >= 5


def test_workbook_fill_juyojiko_kubun() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, KanriHiyou

    wb = Workbook()
    ws = wb.active
    ws.title = "重要事項説明書"
    ws["F7"] = "旧買主"
    ws["F267"] = "旧売主"
    ws["H1116"] = 1700000
    buf = io.BytesIO()
    wb.save(buf)

    ab = Juyojiko(
        bukken_type="区分",
        kainushi=Party(name="株式会社Martial Arts"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="テスト市1番",
                             senyuu=TatemonoHyoji(kaoku_bango="1番の101", meisho="101")),
        horei=HoreiSeigen(kenpei=80, yoseki=400),
        kanri=KanriHiyou(kanrihi_getsugaku=5500, shuzen_getsugaku=4770),
        joken=TorihikiJoken(baibai_daikin=1_700_000),
    )
    for variant in ("37-1", "38-1"):
        bc = transform_ab_to_bc(ab, DEAL)
        sv, sc = cellmaps.build_juyojiko(variant, bc)
        out, n = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
        ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
        assert ws2["F7"].value == "東洋建設ホーム株式会社", variant
        assert ws2["F267"].value == "株式会社Martial Arts", variant
        assert ws2["Q388"].value == 80 and ws2["Q402"].value == 400, variant
        assert ws2["L884"].value == 5500 and ws2["L864"].value == 4770, variant
        assert ws2["H1116"].value == 27_800_000, variant   # BC代金で上書き
        assert n >= 6


def test_juyojiko_checkbox_marks() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen

    # 36-1: 用途地域=第1種住居地域(C364)、区域区分=市街化区域(T331)
    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    ws["C356"] = "■"   # 旧選択（第1種低層）→ □ に戻るはず
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="株式会社Martial Arts"),
                  fudosan=FudosanHyoji(bukken_type="戸建"),
                  horei=HoreiSeigen(kuiki_kubun="市街化区域", yoto="第1種住居地域"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["C364"].value == "■"      # 第1種住居地域 を選択
    assert ws2["C356"].value == "□"      # 旧選択は解除
    assert ws2["T331"].value == "■"      # 市街化区域
    assert ws2["AA331"].value == "□"


def test_juyojiko_chiiki_chiku_checkboxes() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    ws["C370"] = "■"   # 旧: 準防火 → 防火地域選択で □ に
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="株式会社Martial Arts"),
                  fudosan=FudosanHyoji(bukken_type="戸建"),
                  horei=HoreiSeigen(boka="防火地域", nijuni_jo=True, kodo_chiku="第3種高度地区"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["C368"].value == "■"   # 防火地域
    assert ws2["C370"].value == "□"   # 準防火は解除
    assert ws2["C374"].value == "■"   # 建築基準法22条
    assert ws2["C376"].value == "■"   # 高度地区


def test_date_split_fill() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    assert cellmaps._split_wareki("令和7年1月1日") == (7, 1, 1)
    assert cellmaps._split_wareki("2025年4月10日") == (7, 4, 10)
    assert cellmaps._split_wareki("2025-12-01") == (7, 12, 1)
    assert cellmaps._split_wareki(None) is None

    wb = Workbook(); ws = wb.active; ws.title = "不動産売買契約書"
    ws["S59"] = 9   # 旧年（クリア＆上書き）
    buf = io.BytesIO(); wb.save(buf)
    bc = transform_keiyaku_ab_to_bc(
        AB_KEIYAKU,
        {**DEAL, "bc_zankin_date": "2025年12月1日", "bc_loan_shonin_date": "令和7年11月20日"},
    )
    sv, sc = cellmaps.build_keiyaku("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["不動産売買契約書"]
    assert (ws2["S59"].value, ws2["W59"].value, ws2["AA59"].value) == (7, 12, 1)
    assert (ws2["O71"].value, ws2["S71"].value, ws2["W71"].value) == (7, 11, 20)


def test_bundle_merge_pdfs() -> None:
    import bundle
    from pypdf import PdfWriter

    def _pdf(n: int) -> bytes:
        w = PdfWriter()
        for _ in range(n):
            w.add_blank_page(width=200, height=200)
        b = io.BytesIO(); w.write(b)
        return b.getvalue()

    merged, pages = bundle.merge_pdfs([_pdf(2), _pdf(3), _pdf(1)])
    assert pages == 6
    from pypdf import PdfReader
    assert len(PdfReader(io.BytesIO(merged)).pages) == 6


def test_generate_package_both_sheets() -> None:
    from fastapi.testclient import TestClient
    import base64
    import bc_service
    from openpyxl import Workbook

    # 両シートを持つ合成ワークブック（旧データ入り）
    wb = Workbook()
    j = wb.active; j.title = "重要事項説明書"
    j["F7"] = "旧買主"; j["F263"] = "旧売主"
    k = wb.create_sheet("不動産売買契約書")
    k["E123"] = "旧売主"; k["AB123"] = "旧買主"
    buf = io.BytesIO(); wb.save(buf)
    tb = base64.b64encode(buf.getvalue()).decode()

    ab_j = {"bukken_type": "戸建", "kainushi": {"name": "Martial"},
            "fudosan": {"bukken_type": "戸建", "tochi": {"shozai": "x"}},
            "horei": {"kenpei": 60, "yoseki": 200}}
    ab_k = {"bukken_type": "戸建",
            "fudosan": {"bukken_type": "戸建", "tochi": {"shozai": "x"}, "tatemono": {}},
            "daikin": {"baibai_daikin": 23300000, "tetsuke": 300000}}
    c = TestClient(bc_service.app)
    r = c.post("/generate", json={"doc_type": "package", "template": "36-1",
                                  "template_base64": tb, "ab": ab_j, "ab_keiyaku": ab_k,
                                  "deal_master": DEAL})
    assert r.status_code == 200, r.text
    out = base64.b64decode(r.json()["xlsx_base64"])
    wb2 = load_workbook(io.BytesIO(out))
    assert wb2["重要事項説明書"]["F7"].value == "東洋建設ホーム株式会社"   # 重説 買主C
    assert wb2["不動産売買契約書"]["E123"].value == "株式会社Martial Arts"  # 契約 売主B
    assert wb2["不動産売買契約書"]["AE45"].value == 27_800_000             # 契約 BC代金


def test_chiban_split_fill() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    assert cellmaps._split_chiban("○○市○○町12番5") == ("○○市○○町", "12", "5")
    assert cellmaps._split_chiban("△市字横田551番地1") == ("△市字横田", "551", "1")
    assert cellmaps._split_chiban("□市5番") == ("□市", "5", None)
    assert cellmaps._split_chiban("番地なし町") == ("番地なし町", None, None)

    wb = Workbook(); ws = wb.active; ws.title = "不動産売買契約書"
    ws["X11"] = "旧番"; ws["AC11"] = "旧番地"
    buf = io.BytesIO(); wb.save(buf)
    ab = Keiyakusho(
        bukken_type="戸建", urinushi=Party(name="A"), kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="戸建",
                             tochi=TochiHyoji(shozai="ひたちなか市津田東一丁目12番5"),
                             tatemono=TatemonoHyoji()),
        daikin=KeiyakuDaikin(baibai_daikin=23300000, tetsuke=300000),
    )
    bc = transform_keiyaku_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_keiyaku("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["不動産売買契約書"]
    assert ws2["F11"].value == "ひたちなか市津田東一丁目"
    assert ws2["X11"].value == "12" and ws2["AC11"].value == "5"


def test_approval_decision() -> None:
    import approval

    assert approval.decide("✅") == "approve"
    assert approval.decide(":white_check_mark:") == "approve"
    assert approval.decide("❌") == "reject"
    assert approval.decide("x") == "reject"
    assert approval.decide("eyes") == "pending"
    assert approval.decide(None) == "pending"
    assert approval.reaction_from_payload(
        {"event": {"type": "reaction_added", "reaction": "white_check_mark"}}
    ) == "white_check_mark"


def test_approval_endpoint() -> None:
    from fastapi.testclient import TestClient
    import bc_service

    c = TestClient(bc_service.app)
    assert c.post("/approval", json={"type": "url_verification",
                                     "challenge": "abc"}).json() == {"challenge": "abc"}
    assert c.post("/approval", json={"reaction": "✅"}).json()["approved"] is True
    assert c.post("/approval", json={"reaction": "❌"}).json()["decision"] == "reject"


def test_horei_master_lists() -> None:
    import horei_master as H

    # 用途地域は14種（指定なしを含む）
    assert len(H.YOTO_OPTIONS) == 14
    assert H.YOTO_OPTIONS[-1] == "用途地域の指定なし"
    # (3)法令は最新版61件。重複なし。
    assert len(H.OTHER_HOREI_LAWS) == 61
    assert len(set(H.OTHER_HOREI_LAWS)) == 61
    # 最新版で追加された法令を含む
    assert "地域における生物の多様性の増進のための活動の促進等に関する法律" in H.OTHER_HOREI_LAWS
    assert "宅地造成及び特定盛土等規制法" in H.OTHER_HOREI_LAWS
    assert "重要土地等調査法" in H.OTHER_HOREI_LAWS
    # 地域地区
    assert "建築基準法第22条区域" in H.CHIIKI_CHIKU


def test_reference_endpoint() -> None:
    from fastapi.testclient import TestClient
    import bc_service

    d = TestClient(bc_service.app).get("/reference").json()
    assert len(d["other_horei"]) == 61 and len(d["yoto"]) == 14


def test_normalize_horei() -> None:
    import horei_master as H

    assert H.normalize_horei("マンション建替え円滑化法") == "マンションの建替え等の円滑化に関する法律"
    assert H.normalize_horei("盛土規制法") == "宅地造成及び特定盛土等規制法"
    assert H.normalize_horei("生物多様性増進法") == \
        "地域における生物の多様性の増進のための活動の促進等に関する法律"
    assert H.normalize_horei("古都保存法") == "古都保存法"      # 既に正式
    assert H.normalize_horei("未知の条例") == "未知の条例"      # 未知は素通し


def test_extract_normalization() -> None:
    import bc_service

    out = bc_service._normalize_extracted(
        {"horei": {"yoto": "第1種中高層",
                   "other_horei": ["マンション建替え円滑化法", "盛土規制法"]}})
    assert out["horei"]["yoto"] == "第1種中高層住居専用地域"
    assert out["horei"]["other_horei"][0] == "マンションの建替え等の円滑化に関する法律"
    assert out["horei"]["other_horei"][1] == "宅地造成及び特定盛土等規制法"


def test_juyojiko_36_1_gyosha_and_extra() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200, shikichi_saitei="100㎡"))
    deal = {**DEAL, "bc_gyosha_shozai": "東京都港区芝1-2-3", "bc_gyosha_daihyo": "長谷川 光",
            "bc_torikiishi_shimei": "山田 太郎",
            "bc_torikiishi_jimusho_shozai": "東京都港区芝1-2-3 ビル5F"}
    bc = transform_ab_to_bc(ab, deal)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["H23"].value == "東京都港区芝1-2-3"      # 売主側業者の事務所（左欄）
    assert ws2["H31"].value == "長谷川 光"             # 売主側業者の代表者（左欄）
    assert ws2["AF31"].value is None                   # 媒介欄（右）は媒介業者が無ければ空
    assert ws2["H35"].value == "山田 太郎"              # 取引士氏名（左欄）
    assert ws2["H39"].value == "東京都港区芝1-2-3 ビル5F"  # 取引士事務所（左欄）
    assert ws2["R406"].value == "100㎡"                 # 敷地面積最低限度


def test_cover_broker_two_blocks_and_residue_clear() -> None:
    """表紙: 左=売主側業者B・右=媒介業者を両ブロック埋め、媒介が無ければ右はクリア。"""
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    # 記入済みテンプレ想定: 右欄に旧案件（柴崎建設）の残渣を仕込む
    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    for c, v in {"AF21": "埼玉県知事", "AF29": "柴崎建設株式会社",
                 "AF31": "小林 真紀", "AF35": "小林 真紀"}.items():
        ws[c] = v
    buf = io.BytesIO(); wb.save(buf)

    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")))
    deal = {
        **DEAL,
        # 左欄＝売主業者B
        "bc_gyosha_menkyo_no": "東京都知事(1)第105715号",
        "bc_gyosha_shomei": "株式会社Martial Arts",
        "bc_gyosha_daihyo": "長谷川 光",
        "bc_gyosha_tel": "03-6908-2680",
        "bc_torikiishi_toroku_no": "（埼玉）第070441号",
        "bc_torikiishi_shimei": "小玉 浩之",
        # 右欄＝媒介業者
        "bc_baikai_gyosha_menkyo_no": "埼玉県知事(1)第025224号",
        "bc_baikai_gyosha_shomei": "東洋建設ホーム株式会社",
        "bc_baikai_gyosha_daihyo": "石井 光靜",
        "bc_baikai_gyosha_tel": "042-000-2842",
        "bc_baikai_gyosha_is_kyokai_member": True,
        "bc_baikai_gyosha_hosho_kyokai": "公益財団法人全国宅地建物取引業保証協会",
        "bc_baikai_gyosha_hosho_honbu": "公益財団法人全国宅地建物取引業保証協会　埼玉本部",
        "bc_baikai_torikiishi_toroku_no": "第049759号",
        "bc_baikai_torikiishi_shimei": "浅岡 伸之",
    }
    bc = transform_ab_to_bc(ab, deal)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]

    # 左欄＝売主業者B（免許/電話/登録番号の分解差込）
    assert ws2["H21"].value == "東京都知事"
    assert ws2["O21"].value == "(1)"
    assert ws2["S21"].value == "105715"
    assert ws2["H29"].value == "株式会社Martial Arts"
    assert ws2["H31"].value == "長谷川 光"
    assert (ws2["H27"].value, ws2["N27"].value, ws2["T27"].value) == ("03", "6908", "2680")
    assert ws2["H33"].value == "（埼玉）"
    assert ws2["R33"].value == "070441"
    assert ws2["H35"].value == "小玉 浩之"

    # 右欄＝媒介業者（柴崎建設の残渣を上書き）
    assert ws2["AF21"].value == "埼玉県知事"
    assert ws2["AM21"].value == "(1)"
    assert ws2["AQ21"].value == "025224"
    assert ws2["AF29"].value == "東洋建設ホーム株式会社"   # 旧"柴崎建設"を上書き
    assert ws2["AF31"].value == "石井 光靜"               # 旧"小林 真紀"を上書き
    assert ws2["AF35"].value == "浅岡 伸之"               # 旧"小林 真紀"を上書き
    assert ws2["AP33"].value == "049759"
    assert ws2["AA45"].value == "■"                       # 保証協会の社員
    assert ws2["AH47"].value == "公益財団法人全国宅地建物取引業保証協会"
    assert "埼玉本部" in ws2["AH51"].value


def test_keiyaku_house_defaults_and_validate() -> None:
    """BC契約書: 違約金20%・特約は重説参照。validate が漏れ/不整合を検出。"""
    import validate
    from keiyaku_schema import Keiyakusho, KeiyakuDaikin

    ab = Keiyakusho(bukken_type="戸建",
                    daikin=KeiyakuDaikin(baibai_daikin=16900000, iyakukin_wariai=10))
    bc = transform_keiyaku_ab_to_bc(ab, {
        "buyer_C": "高野橋 拓巳", "bc_baibai_daikin": 19800000,
        "bc_tetsuke": 100000, "bc_zankin": 19700000,
        "bc_zankin_date": "令和7年4月30日", "bc_seisan_kisanbi": "令和7年1月1日"})
    assert bc.daikin.iyakukin_wariai == 20                  # 御社標準に上書き
    assert any("重要事項説明書に準拠" in t for t in bc.tokuyaku)  # 特約は重説参照に集約
    issues = validate.validate_keiyaku(bc)
    assert [i for i in issues if i["level"] == "error"] == []

    bad = transform_keiyaku_ab_to_bc(Keiyakusho(bukken_type="戸建"), {})
    fields = {i["field"] for i in validate.validate_keiyaku(bad) if i["level"] == "error"}
    assert "買主C" in fields and "売買代金" in fields


def test_validate_zankin_mismatch() -> None:
    """残代金＞売買代金（実書類の記入ミス相当）を error 検出。"""
    import validate
    from keiyaku_schema import Keiyakusho, KeiyakuDaikin, Party

    bc = Keiyakusho(bukken_type="戸建",
                    urinushi=Party(name="株式会社Martial Arts"),
                    kainushi=Party(name="買主C"),
                    daikin=KeiyakuDaikin(baibai_daikin=21000000, tetsuke=500000,
                                         zankin=22000000, iyakukin_wariai=20))
    issues = validate.validate_keiyaku(bc)
    zk = [i for i in issues if i["field"] == "残代金"]
    assert zk and zk[0]["level"] == "error"   # 22,000,000 > 21,000,000
    # 正しい残代金（20,500,000）なら指摘なし
    bc.daikin.zankin = 20500000
    assert not [i for i in validate.validate_keiyaku(bc) if i["field"] == "残代金"]


def test_validate_bc_cross_doc_consistency() -> None:
    """重説と契約書で買主C・売買代金が食い違うと error。"""
    import validate
    from keiyaku_schema import Keiyakusho, KeiyakuDaikin

    j = transform_ab_to_bc(
        Juyojiko(bukken_type="戸建",
                 fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x"))),
        {"buyer_C": "Ｃさん", "bc_baibai_daikin": 19800000})
    k = transform_keiyaku_ab_to_bc(
        Keiyakusho(bukken_type="戸建", daikin=KeiyakuDaikin()),
        {"buyer_C": "別人", "bc_baibai_daikin": 20000000})
    errs = {i["field"] for i in validate.validate_bc(juyojiko=j, keiyaku=k)
            if i["level"] == "error"}
    assert "買主C" in errs and "売買代金" in errs


def test_house_style_sanme_and_yonin_defaults() -> None:
    """BC変換で御社標準の三為特約（四者間取引の特約）全文と標準容認事項が付く。"""
    import house_style

    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
                  yonin_jiko=["本物件は当社の固有事項"], tokuyaku=["既存特約A"])
    bc = transform_ab_to_bc(ab, {"buyer_C": "C太郎"})
    toku = "\n".join(bc.tokuyaku)
    yonin = "\n".join(bc.yonin_jiko)
    # 三為特約の全文（タイトル＋6節）が入る
    assert "【四者間取引の特約】" in toku
    assert "他人物売買契約" in toku and "所有権留保" in toku
    assert "既存特約A" in toku                       # AB引継ぎは保持
    # 標準容認事項が付与され、物件固有も残る
    assert "現状有姿売買であるため、契約不適合免責" in yonin
    assert "買主の指定する司法書士" in yonin
    assert "本物件は当社の固有事項" in yonin
    # 既に三為記載があれば二重付与しない
    ab2 = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                   fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
                   tokuyaku=["所有権移転先を指定する旨の特約あり"])
    bc2 = transform_ab_to_bc(ab2, {})
    assert sum(1 for t in bc2.tokuyaku if "四者間取引の特約" in t) == 0
    # マスタ整合
    assert house_style.SELLER_B_MASTER["shomei"] == "株式会社Martial Arts"


def test_webui_route_serves_form() -> None:
    """ブラウザ用UI（GET /）が操作画面のHTMLを返す。"""
    from fastapi.testclient import TestClient
    import bc_service

    r = TestClient(bc_service.app).get("/")
    assert r.status_code == 200
    assert "text/html" in r.headers["content-type"]
    assert "BC自動生成" in r.text
    assert "/extract" in r.text and "/generate" in r.text  # 動線が埋まっている
    assert "/masters" in r.text  # プリセット取得の動線


def test_masters_endpoint() -> None:
    """/masters が御社マスタ（売主業者B・取引士・媒介プリセット）を返す。"""
    from fastapi.testclient import TestClient
    import bc_service

    m = TestClient(bc_service.app).get("/masters").json()
    assert m["seller_b"]["shomei"] == "株式会社Martial Arts"
    names = [g["shomei"] for g in m["baikai_gyosha"]]
    assert "東洋建設ホーム株式会社" in names and "柴崎建設株式会社" in names
    assert [t["shimei"] for t in m["seller_b_torikiishi"]]  # 取引士プリセットあり


def test_fill_workbook_sets_print_fit_to_width() -> None:
    """差し込んだシートは横1ページ収め（fitToWidth=1・固定縮尺解除）になる。"""
    import wb_fill
    from openpyxl import Workbook

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    ws.page_setup.scale = 73  # 固定縮尺（右端が切れる元）
    buf = io.BytesIO(); wb.save(buf)
    out, _ = wb_fill.fill_workbook(buf.getvalue(),
                                   {"重要事項説明書": {"A1": "x"}},
                                   {"重要事項説明書": []})
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2.page_setup.fitToWidth == 1
    assert ws2.page_setup.fitToHeight == 0
    assert ws2.page_setup.scale is None
    assert ws2.sheet_properties.pageSetUpPr.fitToPage is True


def test_cover_right_block_cleared_when_no_baikai() -> None:
    """媒介業者が案件マスタに無いとき、右欄の旧残渣が確実にクリアされる。"""
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    for c, v in {"AF21": "埼玉県知事", "AF29": "柴崎建設株式会社",
                 "AF25": "金田第二ビル２B", "AF41": "金田第二ビル２B",
                 "AF31": "小林 真紀", "AH47": "旧保証協会"}.items():
        ws[c] = v
    buf = io.BytesIO(); wb.save(buf)

    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")))
    deal = {**DEAL, "bc_gyosha_shomei": "株式会社Martial Arts"}  # 媒介なし
    bc = transform_ab_to_bc(ab, deal)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    # 商号/免許/代表/所在地2行目（建物名）/保証協会まで右欄の残渣が全て消える
    for c in ("AF21", "AF29", "AF25", "AF41", "AF31", "AH47"):
        assert ws2[c].value is None, f"右欄 {c} の残渣が消えていない"
    assert ws2["H29"].value == "株式会社Martial Arts"   # 左欄は埋まる


def test_juyojiko_36_1_doro_fuzoku() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35"),
                                       fuzoku_tatemono="無"),
                  horei=HoreiSeigen(kenpei=60, yoseki=200, doro_hoko="南西",
                                    doro_haba="約16.00", doro_setsudo="約17.78",
                                    doro="市道番号：2級12号線"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["AK238"].value == "無"           # 附属建物
    assert ws2["D440"].value == "南西"          # 接面道路 方向
    assert ws2["X440"].value == "約16.00"       # 幅員
    assert ws2["AE440"].value == "約17.78"      # 接道長さ
    assert ws2["AL438"].value == "市道番号：2級12号線"  # 備考


def test_juyojiko_36_1_setsubi() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Setsubi

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    ws["G645"] = "■"   # 旧: 私営水道 → 公営選択で □ に
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  setsubi_detail=Setsubi(suidou="公営水道", gas="個別プロパン",
                                         osui="公共下水", zassui="浸透式",
                                         denryoku="東京電力", biko="前面道路に配管あり"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["G643"].value == "■" and ws2["G645"].value == "□"   # 公営水道
    assert ws2["G660"].value == "■"      # 個別プロパン
    assert ws2["G664"].value == "■"      # 汚水 公共下水
    assert ws2["G682"].value == "■"      # 雑排水 浸透式
    assert ws2["G652"].value == "東京電力"
    assert ws2["B695"].value == "前面道路に配管あり"


def test_juyojiko_36_1_saigai() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Saigai

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  saigai=Saigai(zosei_bosai=False, dosha_keikai=True,
                                tsunami_keikai=False, taishin_shindan=False,
                                sekimen_kiroku=False))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["Z795"].value == "■" and ws2["AD795"].value == "□"   # 造成 外
    assert ws2["Z800"].value == "□" and ws2["AD800"].value == "■"   # 土砂警戒 内
    assert ws2["Z807"].value == "■"                                  # 津波 外
    assert ws2["V844"].value == "■" and ws2["R844"].value == "□"     # 耐震 無
    assert ws2["F831"].value == "□"                                  # 石綿記録 無


def test_juyojiko_36_1_hazard_kakunin() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Saigai, Kakunin

    assert cellmaps._split_era_date("平成19年8月10日") == ("平成", 19, 8, 10)
    assert cellmaps._split_era_date("2025-12-01") == ("令和", 7, 12, 1)

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  saigai=Saigai(kozui=True, naisui=False, takashio=False),
                  kakunin=Kakunin(kenchiku_bango="第07UDIIC0345",
                                  kenchiku_date="平成19年8月10日"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["W814"].value == "■" and ws2["AA814"].value == "□"   # 洪水 有
    assert ws2["AM814"].value == "□" and ws2["AQ814"].value == "■"   # 内水 無
    assert ws2["B780"].value == "■" and ws2["AG780"].value == "第07UDIIC0345"
    assert (ws2["R780"].value, ws2["U780"].value, ws2["Y780"].value,
            ws2["AC780"].value) == ("平成", 19, 8, 10)


def test_juyojiko_36_1_touki() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Touki

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建",
                                       tochi=TochiHyoji(shozai="牛久市南7丁目53番35")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  senyuusha_uchi="第三者の占有なし",
                  touki=Touki(tochi_shoyusha_jusho="東京都港区A1-1",
                              tochi_shoyusha_shimei="所有者A",
                              tochi_otsuku="抵当権設定 令和7年2月",
                              tatemono_otsuku="余白"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["L288"].value == "東京都港区A1-1"
    assert ws2["L290"].value == "所有者A"
    assert ws2["L296"].value == "抵当権設定 令和7年2月"
    assert ws2["L316"].value == "余白"
    assert ws2["F277"].value == "第三者の占有なし"


def test_juyojiko_kubun_gyosha_saigai() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Saigai

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="テスト市1番"),
                  horei=HoreiSeigen(kenpei=80, yoseki=400),
                  saigai=Saigai(dosha_keikai=True, tsunami_keikai=False))
    deal = {**DEAL, "bc_gyosha_shozai": "東京都港区芝1-2", "bc_gyosha_daihyo": "長谷川 光",
            "bc_torikiishi_shimei": "山田 太郎", "bc_torikiishi_jimusho_shozai": "港区芝1-2 5F"}
    for variant in ("37-1", "38-1"):
        bc = transform_ab_to_bc(ab, deal)
        sv, sc = cellmaps.build_juyojiko(variant, bc)
        out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
        ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
        assert ws2["H23"].value == "東京都港区芝1-2", variant
        assert ws2["H31"].value == "長谷川 光", variant
        assert ws2["H35"].value == "山田 太郎", variant
        assert ws2["Z1048"].value == "□" and ws2["AD1048"].value == "■", variant  # 土砂内
        assert ws2["Z1055"].value == "■", variant                                   # 津波外


def test_juyojiko_kubun_setsubi_kakunin() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Setsubi, Kakunin, TatemonoHyoji

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="テスト市1番",
                                       senyuu=TatemonoHyoji(kaoku_bango="1番")),
                  horei=HoreiSeigen(kenpei=80, yoseki=400),
                  setsubi_detail=Setsubi(suidou="公営水道", gas="都市ガス",
                                         osui="公共下水", zassui="浸透式", denryoku="東北電力"),
                  kakunin=Kakunin(kenchiku_bango="第ABC123", kenchiku_date="平成2年12月3日"))
    for variant in ("37-1", "38-1"):
        bc = transform_ab_to_bc(ab, DEAL)
        sv, sc = cellmaps.build_juyojiko(variant, bc)
        out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
        ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
        assert ws2["G647"].value == "■" and ws2["G660"].value == "■", variant  # 公営水道/都市ガス
        assert ws2["G686"].value == "■" and ws2["G653"].value == "東北電力", variant
        assert ws2["B1028"].value == "■" and ws2["AG1028"].value == "第ABC123", variant
        assert (ws2["R1028"].value, ws2["U1028"].value) == ("平成", 2), variant


def test_juyojiko_kubun_touki_hazard() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Saigai, Touki, TatemonoHyoji

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="テスト市1番",
                                       senyuu=TatemonoHyoji(kaoku_bango="1番")),
                  horei=HoreiSeigen(kenpei=80, yoseki=400),
                  saigai=Saigai(kozui=True, takashio=False),
                  senyuusha_uchi="賃借中",
                  touki=Touki(tatemono_shoyusha_jusho="福島県A", tatemono_shoyusha_shimei="所有者B",
                              tochi_otsuku="敷地権につき建物と一体"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("37-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["L292"].value == "福島県A" and ws2["L294"].value == "所有者B"
    assert ws2["L320"].value == "敷地権につき建物と一体"
    assert ws2["F281"].value == "賃借中"
    assert ws2["W1062"].value == "■" and ws2["W1064"].value == "□"


def test_aux_sheets_header() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active; ws.title = "335.取引完了確認書"
    wb.create_sheet("735-1.領収書")
    buf = io.BytesIO(); wb.save(buf)

    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")))
    bc = transform_ab_to_bc(ab, DEAL)
    av, ac = cellmaps.build_aux(bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), av, ac)
    wb2 = load_workbook(io.BytesIO(out))
    assert wb2["335.取引完了確認書"]["G33"].value == "株式会社Martial Arts"
    assert wb2["335.取引完了確認書"]["AB33"].value == "東洋建設ホーム株式会社"
    assert wb2["735-1.領収書"]["G21"].value == "東洋建設ホーム株式会社"


def test_juyojiko_biko_freeform() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, TatemonoHyoji

    # 36-1 (B1196)
    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  yonin_jiko=["本物件は現況有姿売買"], tokuyaku=["設備表は交付しない"])
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    v = load_workbook(io.BytesIO(out))["重要事項説明書"]["B1196"].value
    assert "本物件は現況有姿売買" in v and "設備表は交付しない" in v
    assert "四者間取引の特約" in v or "他人物売買" in v   # 御社標準の三為特約全文も含む

    # 区分 (B1449)
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "重要事項説明書"
    buf2 = io.BytesIO(); wb2.save(buf2)
    abk = Juyojiko(bukken_type="区分", kainushi=Party(name="M"),
                   fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x",
                                        senyuu=TatemonoHyoji()),
                   horei=HoreiSeigen(kenpei=80, yoseki=400), yonin_jiko=["集合住宅"])
    bck = transform_ab_to_bc(abk, DEAL)
    svk, sck = cellmaps.build_juyojiko("37-1", bck)
    outk, _ = wb_fill.fill_workbook(buf2.getvalue(), svk, sck)
    # 区分は 容認事項=B1366 / 特約=B1449 に分かれる
    assert "集合住宅" in load_workbook(io.BytesIO(outk))["重要事項説明書"]["B1366"].value


def test_juyojiko_shakuchi_into_biko() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen, Shakuchi

    # 借地物件: 専用借地シートは未照合のため、借地条件をⅤ備考(B1196)へ転記
    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200),
                  shakuchi=Shakuchi(shakuchiken_shurui="普通借地権",
                                    sonzoku_kikan="令和3年〜令和23年",
                                    jidai_kingaku=30_000, jidai_tani="月額",
                                    teichi_shoyusha_shimei="地主 太郎"))
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    v = load_workbook(io.BytesIO(out))["重要事項説明書"]["B1196"].value
    assert "【借地条件】" in v
    assert "普通借地権" in v and "月額30,000円" in v and "地主 太郎" in v


def test_juyojiko_36_1_section_biko() -> None:
    import cellmaps
    import wb_fill
    from openpyxl import Workbook
    from juyojiko_schema import HoreiSeigen

    wb = Workbook(); ws = wb.active; ws.title = "重要事項説明書"
    buf = io.BytesIO(); wb.save(buf)
    ab = Juyojiko(bukken_type="戸建", kainushi=Party(name="M"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="x"),
                                       fuzoku_tatemono_detail="物置 軽量鉄骨造"),
                  horei=HoreiSeigen(kenpei=60, yoseki=200, suigai_shozai="洪水HM参照"),
                  seisan_biko="公租公課は日割り清算")
    bc = transform_ab_to_bc(ab, DEAL)
    sv, sc = cellmaps.build_juyojiko("36-1", bc)
    out, _ = wb_fill.fill_workbook(buf.getvalue(), sv, sc)
    ws2 = load_workbook(io.BytesIO(out))["重要事項説明書"]
    assert ws2["B250"].value == "物置 軽量鉄骨造"
    assert ws2["O818"].value == "洪水HM参照"
    assert ws2["B891"].value == "公租公課は日割り清算"


def test_fidelity_check_tool() -> None:
    # 照合エンジン fidelity_check.compare の最小検証（合成WB・実物件WB不要）
    import os
    import tempfile

    from openpyxl import Workbook

    import fidelity_check as F
    d = tempfile.mkdtemp()
    blank_p = os.path.join(d, "blank.xlsx")
    truth_p = os.path.join(d, "truth.xlsx")
    wb = Workbook(); wb.active.title = "重要事項説明書"; wb.save(blank_p)
    bc = Juyojiko(bukken_type="戸建", kainushi=Party(name="買主C"),
                  urinushi=Party(name="売主B", address="東京都X"),
                  fudosan=FudosanHyoji(bukken_type="戸建", tochi=TochiHyoji(shozai="A市1番2")),
                  horei=HoreiSeigen(kenpei=60, yoseki=200))
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "重要事項説明書"
    ws2["F263"] = "売主B"      # パイプラインと一致させる
    ws2["F7"] = "別の買主"      # パイプライン(買主C)と不一致にする
    wb2.save(truth_p)
    s = F.compare(bc, "36-1", "juyojiko", blank_p, truth_p)
    diff_cells = {c for c, _, _ in s["diffs"]}
    assert s["written"] > 0
    assert "F263" not in diff_cells      # 一致セルは diffs に出ない
    assert "F7" in diff_cells            # 不一致セルは diffs に出る
    assert s["match"] >= 1 and 0.0 <= s["match_rate"] <= 1.0


def test_build_juyojiko_edition_b_remap() -> None:
    # Edition B(37-1): 値セルは列差替、法令格子・建築時期・取引条件(違約金/担保)は+2行
    import cellmaps
    import cellmap_grids
    bc = Juyojiko(
        bukken_type="区分", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x", ittou_meisho="一棟名",
                             senyuu=TatemonoHyoji(meisho="専有名", chikujiki="平成2年12月3日")),
        horei=HoreiSeigen(kenpei=60, yoseki=200, kuiki_kubun="市街化区域",
                          other_horei=["古都保存法"]),
        kanri=KanriHiyou(kanri_kumiai="○○管理組合"),
        joken=TorihikiJoken(iyakukin_wariai=20, tanpo_sekinin="講じない"))
    va = cellmaps.build_juyojiko("37-1", bc, edition="A")[0]["重要事項説明書"]
    vb = cellmaps.build_juyojiko("37-1", bc, edition="B")[0]["重要事項説明書"]
    # A版: 建蔽率Q388・管理組合K900・違約金O1256
    assert va["Q388"] == 60 and va["K900"] == "○○管理組合" and va["O1256"] == "■"
    # B版: 列差替（D388/D900）、A版セルは消える
    assert vb["D388"] == 60 and "Q388" not in vb
    assert vb["D900"] == "○○管理組合" and "K900" not in vb
    assert vb["U207"] == "専有名" and "AL207" not in vb        # 専有建物名称
    # 建築時期 +2行（L213→L215）
    assert vb["L215"] == "平成" and "L213" not in vb
    # 法令格子 +2行（古都保存法）
    a_koto = cellmap_grids.OTHER_HOREI_MARKS["37-1"]["古都保存法"]
    assert va[a_koto] == "■" and vb[cellmaps._shift_row(a_koto, 2)] == "■"
    # 区域区分チェックは同位置（市街化 T335 不変）
    assert va["T335"] == "■" and vb["T335"] == "■"
    # 取引条件(違約金/担保)は +2 行（O1256→O1258 / T1328→T1330）
    assert va["O1256"] == "■" and vb["O1258"] == "■" and "O1256" not in vb
    assert va["W1256"] == 20 and vb["W1258"] == 20
    # 担保「2.講じない」= Z1328→Z1330（講じる T1328→T1330 は □）
    assert va["Z1328"] == "■" and vb["Z1330"] == "■" and "Z1328" not in vb
    # 38-1 は B 写像の対象外（区分でも variant!=37-1）
    v38 = cellmaps.build_juyojiko("38-1", bc, edition="B")[0]["重要事項説明書"]
    assert v38["Q388"] == 60   # 写像されない


def test_edition_b_cover_block_shift() -> None:
    # 表紙の宅建業者・供託所欄は B版で 免許番号行(21)不動・所在地行(23)以降が +2 行。
    # 実B版テンプレでラベルが所在H25/商号H31/代表H33/供託所J49〜と2行下がるため。
    import cellmaps
    from juyojiko_schema import Gyosha
    bc = Juyojiko(
        bukken_type="区分", kainushi=Party(name="M"),
        fudosan=FudosanHyoji(bukken_type="区分", ittou_shozai="x"),
        gyosha=Gyosha(menkyo_no="東京都知事(1)第105715号", shozai="東京都中央区日本橋",
                      shomei="株式会社Martial Arts", daihyo="長谷川 光",
                      is_kyokai_member=True, hosho_kyokai="不動産保証協会"))
    va = cellmaps.build_juyojiko("37-1", bc, edition="A")[0]["重要事項説明書"]
    vb = cellmaps.build_juyojiko("37-1", bc, edition="B")[0]["重要事項説明書"]
    # 免許番号は行21で不動（A/B 同一）
    assert va["S21"] == vb["S21"] == "105715"
    # 所在(H23→H25)・商号(H29→H31)・代表(H31→H33)は +2 行
    assert va["H23"] == "東京都中央区日本橋" and vb["H25"] == "東京都中央区日本橋"
    assert va["H29"] == "株式会社Martial Arts" and vb["H31"] == "株式会社Martial Arts"
    assert va["H31"] == "長谷川 光" and vb["H33"] == "長谷川 光"
    # A版の商号セル(H29)にはB版で商号が残らない（+2で H31 へ移動）
    assert "H23" not in vb and vb.get("H29") != "株式会社Martial Arts"
    # 保証協会員チェック(C45→C47)・協会名(J47→J49)も +2 行
    assert va["C45"] == "■" and vb["C47"] == "■"
    assert va["J47"] == "不動産保証協会" and vb["J49"] == "不動産保証協会"


def test_detect_kubun_edition() -> None:
    # 指定建蔽率ラベルの行で区分様式の版を判定（A=388 / B=390）
    import cellmaps
    from openpyxl import Workbook
    for row, expect in ((388, "A"), (390, "B"), (392, "unknown")):
        ws = Workbook().active
        ws.cell(row, 11).value = "指定建蔽率"   # K列付近
        assert cellmaps.detect_kubun_edition(ws) == expect
    # ラベルが無ければ unknown
    assert cellmaps.detect_kubun_edition(Workbook().active) == "unknown"


def test_auth_password_and_session(tmp_path, monkeypatch) -> None:
    # パスワードハッシュ・検証、HMAC署名セッションの往復・改ざん検知
    monkeypatch.setenv("BC_USERS_FILE", str(tmp_path / "users.json"))
    monkeypatch.setenv("BC_SESSION_SECRET", "test-secret")
    import importlib
    import auth
    importlib.reload(auth)
    h = auth.hash_password("correct horse")
    assert auth.verify_password("correct horse", h)
    assert not auth.verify_password("wrong", h)
    auth.save_users({"u1": {"pw_hash": h, "display_name": "User1"}})
    assert auth.authenticate("u1", "correct horse")
    assert not auth.authenticate("u1", "nope")
    assert not auth.authenticate("ghost", "x")
    tok = auth.create_session("u1")
    assert auth.verify_session(tok) == "u1"
    assert auth.verify_session(tok + "x") is None        # 署名改ざん
    assert auth.verify_session("garbage") is None
    # 台帳から消えたユーザーのトークンは無効化される
    auth.save_users({})
    assert auth.verify_session(tok) is None


def test_auth_disabled_is_backward_compatible(tmp_path, monkeypatch) -> None:
    # ユーザー未登録なら認証は無効＝従来どおり全エンドポイントが開いている
    monkeypatch.setenv("BC_USERS_FILE", str(tmp_path / "none.json"))
    import importlib
    import auth
    import bc_service
    importlib.reload(auth)
    importlib.reload(bc_service)
    from fastapi.testclient import TestClient
    c = TestClient(bc_service.app)
    assert c.get("/").status_code == 200
    assert c.get("/masters").status_code == 200
    assert c.get("/me").json()["auth_enabled"] is False


def test_auth_enabled_gates_endpoints(tmp_path, monkeypatch) -> None:
    # ユーザー登録で認証必須。未ログインAPIは401・ブラウザGETはログインへ、
    # ログイン成功でクッキー発行→アクセス可、ログアウトで再び遮断。/health は常時開放。
    monkeypatch.setenv("BC_USERS_FILE", str(tmp_path / "users.json"))
    monkeypatch.setenv("BC_SESSION_SECRET", "test-secret")
    import importlib
    import auth
    import bc_service
    importlib.reload(auth)
    importlib.reload(bc_service)
    auth.save_users({"hikaru": {"pw_hash": auth.hash_password("pw12345678"),
                                "display_name": "長谷川"}})
    from fastapi.testclient import TestClient
    c = TestClient(bc_service.app)
    assert c.get("/masters").status_code == 401
    assert c.get("/health").status_code == 200          # 死活監視は常時開放
    r = c.get("/", headers={"accept": "text/html"}, follow_redirects=False)
    assert r.status_code == 303 and "/login" in r.headers["location"]
    # 誤パスワードはログインへ差し戻し
    r = c.post("/login", data={"username": "hikaru", "password": "bad"},
               follow_redirects=False)
    assert r.status_code == 303 and "error" in r.headers["location"]
    # 正しいログイン→クッキー→アクセス可
    r = c.post("/login", data={"username": "hikaru", "password": "pw12345678"},
               follow_redirects=False)
    assert r.status_code == 303 and auth.COOKIE_NAME in r.cookies
    assert c.get("/masters").status_code == 200
    assert c.get("/me").json()["display_name"] == "長谷川"
    c.get("/logout", follow_redirects=False)
    assert c.get("/masters").status_code == 401


def test_extract_json_and_merge_helpers() -> None:
    import bc_service
    assert bc_service._parse_json_loose('前 {"a":1} 後') == {"a": 1}
    assert bc_service._parse_json_loose('```json\n{"b":2}\n```') == {"b": 2}
    assert bc_service._parse_json_loose("no json") is None
    merged = bc_service._merge_extracted(
        {"x": "A", "list": [1], "d": {"p": ""}},
        {"x": "B", "y": "Y", "list": [1, 2], "d": {"p": "q"}})
    assert merged == {"x": "A", "list": [1, 2], "d": {"p": "q"}, "y": "Y"}


def test_extract_never_500(monkeypatch) -> None:
    # どんな入力でも /extract は 200。読めなければ空データ＋警告（手入力で続行可能）。
    import bc_service
    from fastapi.testclient import TestClient
    c = TestClient(bc_service.app)
    # 1) APIキー未設定 → 200＋警告
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    r = c.post("/extract", json={"doc_type": "juyojiko", "text": "本文"})
    assert r.status_code == 200 and r.json()["extracted"] == {}
    assert "手入力" in r.json()["warning"]
    # 2) 資料もテキストも無い → 200（例外にしない）
    assert c.post("/extract", json={"doc_type": "juyojiko"}).status_code == 200
    # 3) API呼び出しが例外を投げても 200＋警告（500にしない）
    monkeypatch.setenv("ANTHROPIC_API_KEY", "dummy")
    monkeypatch.setattr(bc_service, "_call_claude_json",
                        lambda dt, pieces: (_ for _ in ()).throw(RuntimeError("boom")))
    r = c.post("/extract", json={"doc_type": "juyojiko", "text": "本文"})
    assert r.status_code == 200 and r.json()["extracted"] == {}
    assert "手入力" in r.json()["warning"]
    # 4) 成功時は素通り（警告なし）
    monkeypatch.setattr(bc_service, "_call_claude_json", lambda dt, pieces: {"tokuyaku": ["x"]})
    r = c.post("/extract", json={"doc_type": "juyojiko", "text": "本文"})
    assert r.status_code == 200 and r.json()["extracted"].get("tokuyaku") == ["x"]
    assert r.json()["warning"] == ""


def test_login_next_sanitized_and_favicon(tmp_path, monkeypatch) -> None:
    # ログイン画面の next は反射XSS/オープンリダイレクトを防ぐ。favicon は204。
    monkeypatch.setenv("BC_USERS_FILE", str(tmp_path / "users.json"))
    monkeypatch.setenv("BC_SESSION_SECRET", "x")
    import importlib
    import auth
    import bc_service
    importlib.reload(auth)
    importlib.reload(bc_service)
    auth.save_users({"u": {"pw_hash": auth.hash_password("pw12345678")}})
    from fastapi.testclient import TestClient
    c = TestClient(bc_service.app)
    r = c.get("/login", params={"next": '"><script>alert(1)</script>'})
    assert "<script>alert" not in r.text and 'value="/"' in r.text
    assert 'value="/"' in c.get("/login", params={"next": "//evil.com"}).text
    assert 'value="/generate"' in c.get("/login", params={"next": "/generate"}).text
    assert c.get("/favicon.ico").status_code == 204


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"PASS {fn.__name__}")
    print(f"\n{len(fns)} tests passed")
